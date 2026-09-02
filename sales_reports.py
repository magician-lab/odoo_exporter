# =========================================================
# Sales Reports Service - standalone Flask app on port 5050
# =========================================================
# Independent from app.py (separate process, own templates).
# Reads the same Survitec Odoo 19.2 SaaS instance over XML-RPC.
#
# ODOO 19 SCHEMA FACTS (verified against live DB - do not break):
#   - pos.order.line has NO price_total in this build; revenue is
#     price_subtotal. Only state in ('paid','done') counts.
#   - sale.order.line uses product_uom_qty + price_total; only
#     state = 'sale' counts.
#   - product_id on both line models points at product.product
#     (variant); group to the template via product_tmpl_id.
#   - date_order arrives as an XML-RPC datetime; str() -> "YYYY-MM-DD HH:MM:SS".
# =========================================================

from flask import Flask, jsonify, render_template, request, Response, send_file
import xmlrpc.client
import os
import time
import csv
from io import StringIO
from dotenv import load_dotenv
from datetime import datetime
from collections import defaultdict
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

ODOO_URL = os.getenv("ODOO_URL")
ODOO_DB = os.getenv("ODOO_DB")
ODOO_USERNAME = os.getenv("ODOO_USERNAME")
ODOO_PASSWORD = os.getenv("ODOO_API_KEY") or os.getenv("ODOO_PASSWORD")

app = Flask(__name__, template_folder="templates/sales")

common = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/common", allow_none=True)
uid = common.authenticate(ODOO_DB, ODOO_USERNAME, ODOO_PASSWORD, {})
if not uid:
    raise SystemExit("AUTH FAILED - check .env credentials")

models = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/object", allow_none=True)

_CACHE = {"ts": 0.0, "data": None}
CACHE_TTL = 300.0
_FIELDS_CACHE = {}

POS_OK_STATES = ("paid", "done")
SO_OK_STATES = ("sale",)


def clear_cache():
    _FIELDS_CACHE.clear()
    _CACHE["ts"] = 0.0
    _CACHE["data"] = None


def safe_call(model, method, args=None, kwargs=None, retries=3):
    if args is None:
        args = []
    if kwargs is None:
        kwargs = {}
    last_err = None
    for attempt in range(retries):
        try:
            return models.execute_kw(ODOO_DB, uid, ODOO_PASSWORD, model, method, args, kwargs)
        except Exception as e:
            last_err = e
            print(f"\nODOO ERROR | model={model} method={method} attempt={attempt + 1}/{retries} | {e}")
            time.sleep(1 + attempt * 2)
    print(f"\nODOO FAILED PERMANENTLY | model={model} method={method} | {last_err}")
    return []


def get_valid_model_fields(model, preferred_fields):
    if not preferred_fields:
        return []
    if model not in _FIELDS_CACHE:
        try:
            _FIELDS_CACHE[model] = models.execute_kw(
                ODOO_DB, uid, ODOO_PASSWORD, model, "fields_get",
                [], {"attributes": ["name"]},
            )
        except Exception:
            _FIELDS_CACHE[model] = None
    field_map = _FIELDS_CACHE.get(model)
    if not field_map:
        return preferred_fields
    available = set(field_map.keys())
    valid = [f for f in preferred_fields if f in available]
    return valid or preferred_fields


def search_read(model, fields, domain=None, limit=100000):
    return safe_call(model, "search_read", [domain or []], {"fields": fields, "limit": limit})


# =========================================================
# DATA FETCH
# =========================================================

def fetch_sales_lines():
    """All valid POS + web-sale lines with date, channel and product info."""
    now = time.time()
    if _CACHE["data"] and now - _CACHE["ts"] < CACHE_TTL:
        return _CACHE["data"]

    pos_line_fields = get_valid_model_fields("pos.order.line", ["id", "product_id", "qty", "price_subtotal", "order_id"])
    pos_order_fields = get_valid_model_fields("pos.order", ["id", "date_order", "state"])
    so_line_fields = get_valid_model_fields("sale.order.line", ["id", "product_id", "product_uom_qty", "price_total", "order_id"])
    so_order_fields = get_valid_model_fields("sale.order", ["id", "date_order", "state"])
    variant_fields = get_valid_model_fields("product.product", ["id", "display_name", "name", "product_tmpl_id"])
    template_fields = get_valid_model_fields("product.template", ["id", "name", "categ_id"])

    pos_lines = search_read("pos.order.line", pos_line_fields)
    pos_orders = {o["id"]: o for o in search_read("pos.order", pos_order_fields)}
    so_lines = search_read("sale.order.line", so_line_fields)
    so_orders = {o["id"]: o for o in search_read("sale.order", so_order_fields)}
    variants = {v["id"]: v for v in search_read("product.product", variant_fields)}
    templates = {t["id"]: t for t in search_read("product.template", template_fields)}

    lines = []
    counts = {"pos": 0, "so": 0}

    def _fold(lines_raw, orders, ok_states, qty_field, rev_field, channel):
        for line in lines_raw:
            pid_ref = line.get("product_id")
            if not pid_ref:
                continue
            order = None
            oid = line.get("order_id")
            if oid:
                order = orders.get(oid[0])
            if not order or order.get("state") not in ok_states:
                continue
            date_str = str(order.get("date_order") or "")
            if not date_str:
                continue
            variant = variants.get(pid_ref[0]) or {}
            tid_ref = variant.get("product_tmpl_id")
            tid = tid_ref[0] if tid_ref else None
            template = templates.get(tid) or {}
            lines.append({
                "id": line.get("id"),
                "product_id": pid_ref[0],
                "template_id": tid,
                "product_name": variant.get("display_name") or variant.get("name") or "",
                "template_name": template.get("name") or "",
                "category": template.get("categ_id")[1] if template.get("categ_id") else "",
                "qty": float(line.get(qty_field) or 0),
                "revenue": float(line.get(rev_field) or 0),
                "date": date_str[:10],
                "month": date_str[:7],
                "channel": channel,
                "order_id": order["id"],
            })

    _fold(pos_lines, pos_orders, POS_OK_STATES, "qty", "price_subtotal", "POS")
    counts["pos"] = len(pos_lines)
    _fold(so_lines, so_orders, SO_OK_STATES, "product_uom_qty", "price_total", "Sales Order")
    counts["so"] = len(so_lines)

    lines.sort(key=lambda x: x["date"])
    months = []
    seen = set()
    for l in lines:
        if l["month"] not in seen:
            seen.add(l["month"])
            months.append(l["month"])

    result = {
        "lines": lines,
        "months": months,
        "counts": counts,
    }
    _CACHE["ts"] = now
    _CACHE["data"] = result
    return result


# =========================================================
# AGGREGATION
# =========================================================

def _monthly_map(lines):
    monthly = defaultdict(lambda: {"revenue": 0.0, "units": 0.0, "orders": set()})
    for l in lines:
        m = monthly[l["month"]]
        m["revenue"] += l["revenue"]
        m["units"] += l["qty"]
        m["orders"].add((l["channel"], l["order_id"]))
    return monthly


def build_payload(month=None):
    data = fetch_sales_lines()
    lines = data["lines"]
    months_sorted = data["months"]
    latest = months_sorted[-1] if months_sorted else None

    if month is None or (month not in months_sorted and month != "all"):
        month = latest or "all"

    monthly = _monthly_map(lines)
    selected = lines if month == "all" else [l for l in lines if l["month"] == month]

    products = {}
    for l in lines:
        key = l["template_id"] or ("v%d" % l["product_id"])
        p = products.get(key)
        if p is None:
            p = {
                "key": key,
                "template_id": l["template_id"],
                "name": l["template_name"] or l["product_name"],
                "category": l["category"],
                "first_sale": l["date"],
                "last_sale": l["date"],
                "monthly": defaultdict(lambda: {"revenue": 0.0, "units": 0.0}),
                "channels": defaultdict(lambda: {"revenue": 0.0, "units": 0.0}),
                "variants": defaultdict(lambda: {"name": "", "revenue": 0.0, "units": 0.0}),
            }
            products[key] = p
        p["first_sale"] = min(p["first_sale"], l["date"])
        p["last_sale"] = max(p["last_sale"], l["date"])
        b = p["monthly"][l["month"]]
        b["revenue"] += l["revenue"]
        b["units"] += l["qty"]
        ch = p["channels"][l["channel"]]
        ch["revenue"] += l["revenue"]
        ch["units"] += l["qty"]
        v = p["variants"][l["product_id"]]
        v["name"] = l["product_name"]
        v["revenue"] += l["revenue"]
        v["units"] += l["qty"]

    sel_products = defaultdict(lambda: {"revenue": 0.0, "units": 0.0, "orders": set()})
    channel_totals = defaultdict(lambda: {"revenue": 0.0, "units": 0.0, "orders": set()})
    date_lines = defaultdict(lambda: defaultdict(lambda: {"name": "", "units": 0.0, "revenue": 0.0}))
    kpi_orders = set()
    for l in selected:
        key = l["template_id"] or ("v%d" % l["product_id"])
        sp = sel_products[key]
        sp["revenue"] += l["revenue"]
        sp["units"] += l["qty"]
        sp["orders"].add((l["channel"], l["order_id"]))
        ch = channel_totals[l["channel"]]
        ch["revenue"] += l["revenue"]
        ch["units"] += l["qty"]
        ch["orders"].add((l["channel"], l["order_id"]))
        kpi_orders.add((l["channel"], l["order_id"]))
        cell = date_lines[key][(l["date"], l["product_id"])]
        cell["name"] = l["product_name"]
        cell["units"] += l["qty"]
        cell["revenue"] += l["revenue"]

    product_rows = []
    for key, p in products.items():
        sp = sel_products.get(key)
        if not sp or not sp["orders"]:
            continue
        monthly_bd = [
            {"month": m, "revenue": round(d["revenue"], 2), "units": round(d["units"], 2)}
            for m, d in sorted(p["monthly"].items())
        ]
        channel_bd = [
            {"name": n, "revenue": round(d["revenue"], 2), "units": round(d["units"], 2)}
            for n, d in sorted(p["channels"].items(), key=lambda kv: -kv[1]["revenue"])
        ]
        variants = sorted(
            [{"name": v["name"] or vid, "revenue": round(v["revenue"], 2), "units": round(v["units"], 2)}
             for vid, v in p["variants"].items()],
            key=lambda x: -x["revenue"],
        )
        sales_by_date = sorted(
            [{"date": d, "product": cell["name"], "units": round(cell["units"], 2), "revenue": round(cell["revenue"], 2)}
             for (d, _pid), cell in date_lines.get(key, {}).items()],
            key=lambda x: (x["date"], x["product"]),
        )
        product_rows.append({
            "key": key,
            "template_id": p["template_id"],
            "name": p["name"],
            "category": p["category"] or "Uncategorized",
            "revenue": round(sp["revenue"], 2),
            "units": round(sp["units"], 2),
            "orders": len(sp["orders"]),
            "avg_price": round(sp["revenue"] / sp["units"], 2) if sp["units"] else None,
            "first_sale": p["first_sale"],
            "last_sale": p["last_sale"],
            "channel_breakdown": channel_bd,
            "monthly_breakdown": monthly_bd,
            "variants": variants,
            "sales_by_date": sales_by_date,
        })
    product_rows.sort(key=lambda x: -x["revenue"])

    trend = []
    for m in months_sorted:
        d = monthly[m]
        trend.append({"month": m, "revenue": round(d["revenue"], 2), "units": round(d["units"], 2), "orders": len(d["orders"])})

    kpi_revenue = sum(sp["revenue"] for sp in sel_products.values())
    kpi_units = sum(sp["units"] for sp in sel_products.values())
    categories = {p["category"] for p in product_rows}

    return {
        "month": month,
        "months": months_sorted,
        "kpis": {
            "revenue": round(kpi_revenue, 2),
            "units": round(kpi_units, 2),
            "orders": len(kpi_orders),
            "products": len(product_rows),
            "lines": len(selected),
            "categories": len(categories),
            "avg_order_value": round(kpi_revenue / len(kpi_orders), 2) if kpi_orders else 0.0,
        },
        "channels": [
            {"name": n, "revenue": round(d["revenue"], 2), "units": round(d["units"], 2), "orders": len(d["orders"])}
            for n, d in sorted(channel_totals.items(), key=lambda kv: -kv[1]["revenue"])
        ],
        "monthly_trend": trend,
        "products": product_rows,
        "fetch_health": {"lines": len(lines), "pos_lines": data["counts"]["pos"], "so_lines": data["counts"]["so"]},
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }


# =========================================================
# ROUTES
# =========================================================

@app.route("/")
def home():
    return render_template("report.html")


@app.route("/api/sales")
def api_sales():
    if request.args.get("refresh") == "1":
        clear_cache()
    return jsonify(build_payload(request.args.get("month")))


@app.route("/api/export")
def api_export():
    payload = build_payload(request.args.get("month"))
    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Month", "Product", "Category", "Units", "Revenue (KES)", "Avg Price (KES)", "Orders", "First Sale", "Last Sale"])
    for p in payload["products"]:
        writer.writerow([
            payload["month"],
            p["name"],
            p["category"],
            p["units"],
            p["revenue"],
            p["avg_price"] if p["avg_price"] is not None else "",
            p["orders"],
            p["first_sale"],
            p["last_sale"],
        ])
    month_tag = payload["month"] if payload["month"] != "all" else "all_months"
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=sales_{month_tag}.csv"},
    )


# =========================================================
# EXCEL EXPORT (date-range, grouped by product)
# =========================================================

def build_from_date_report(from_date):
    data = fetch_sales_lines()
    selected = [l for l in data["lines"] if l["date"] >= from_date]

    groups = defaultdict(lambda: {
        "product": "", "template": "", "category": "",
        "units": 0.0, "revenue": 0.0, "orders": set(), "lines": 0,
        "first": None, "last": None,
    })
    for l in selected:
        g = groups[l["product_name"] or l["template_name"]]
        g["product"] = l["product_name"] or l["template_name"] or f"Product {l['product_id']}"
        g["template"] = l["template_name"]
        g["category"] = l["category"] or "Uncategorized"
        g["units"] += l["qty"]
        g["revenue"] += l["revenue"]
        g["orders"].add((l["channel"], l["order_id"]))
        g["lines"] += 1
        g["first"] = l["date"] if g["first"] is None or l["date"] < g["first"] else g["first"]
        g["last"] = l["date"] if g["last"] is None or l["date"] > g["last"] else g["last"]

    total_units = sum(g["units"] for g in groups.values())
    total_revenue = sum(g["revenue"] for g in groups.values())
    order_keys = set()
    for g in groups.values():
        order_keys |= g["orders"]

    summary = []
    for g in groups.values():
        summary.append({
            "product": g["product"],
            "template": g["template"],
            "category": g["category"],
            "units": round(g["units"], 2),
            "revenue": round(g["revenue"], 2),
            "avg_price": round(g["revenue"] / g["units"], 2) if g["units"] else None,
            "orders": len(g["orders"]),
            "lines": g["lines"],
            "first": g["first"],
            "last": g["last"],
            "unit_share": round(g["units"] / total_units * 100, 1) if total_units else 0.0,
            "rev_share": round(g["revenue"] / total_revenue * 100, 1) if total_revenue else 0.0,
        })
    summary.sort(key=lambda x: (-x["revenue"], -x["units"]))

    detail = [{
        "date": l["date"],
        "product": l["product_name"] or l["template_name"] or f"Product {l['product_id']}",
        "template": l["template_name"],
        "category": l["category"] or "Uncategorized",
        "channel": l["channel"],
        "order_id": l["order_id"],
        "units": round(l["qty"], 2),
        "revenue": round(l["revenue"], 2),
        "unit_price": round(l["revenue"] / l["qty"], 2) if l["qty"] else None,
    } for l in sorted(selected, key=lambda x: x["date"])]

    return {
        "from_date": from_date,
        "to_date": datetime.now().strftime("%Y-%m-%d"),
        "summary": summary,
        "detail": detail,
        "totals": {
            "units": round(total_units, 2),
            "revenue": round(total_revenue, 2),
            "orders": len(order_keys),
            "lines": len(selected),
            "products": len(summary),
        },
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }


def write_from_date_excel(report):
    os.makedirs("exports", exist_ok=True)
    filename = f"exports/sales_from_{report['from_date']}_{datetime.now().strftime('%H%M%S')}.xlsx"

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sum_df = pd.DataFrame([{
        "#": i + 1,
        "Product": s["product"],
        "Template": s["template"],
        "Category": s["category"],
        "Units Sold": s["units"],
        "Revenue (KES)": s["revenue"],
        "Avg Price (KES)": s["avg_price"] if s["avg_price"] is not None else "",
        "Orders": s["orders"],
        "Sale Lines": s["lines"],
        "First Sale": s["first"],
        "Last Sale": s["last"],
        "% of Units": s["unit_share"],
        "% of Revenue": s["rev_share"],
    } for i, s in enumerate(report["summary"])])

    det_df = pd.DataFrame([{
        "Date": d["date"],
        "Product": d["product"],
        "Template": d["template"],
        "Category": d["category"],
        "Channel": d["channel"],
        "Order ID": d["order_id"],
        "Units": d["units"],
        "Revenue (KES)": d["revenue"],
        "Unit Price (KES)": d["unit_price"] if d["unit_price"] is not None else "",
    } for d in report["detail"]])

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        info = {
            "Period From": report["from_date"],
            "Period To": report["to_date"],
            "Total Units Sold": report["totals"]["units"],
            "Total Revenue (KES)": report["totals"]["revenue"],
            "Total Orders": report["totals"]["orders"],
            "Total Sale Lines": report["totals"]["lines"],
            "Distinct Products": report["totals"]["products"],
            "Generated At": report["generated_at"],
            "Source": "Survitec Odoo SaaS 19.2 (POS paid/done + sale.order sale)",
        }
        pd.DataFrame(list(info.items()), columns=["Key", "Value"]).to_excel(writer, index=False, sheet_name="Info")
        ws_info = writer.sheets["Info"]
        ws_info.column_dimensions["A"].width = 22
        ws_info.column_dimensions["B"].width = 60
        for cell in ws_info[1]:
            cell.font = Font(bold=True)

        sum_df.to_excel(writer, index=False, sheet_name="By Product")
        ws = writer.sheets["By Product"]
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[letter].width = min(max_len + 5, 60)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        det_df.to_excel(writer, index=False, sheet_name="Sales Detail")
        ws2 = writer.sheets["Sales Detail"]
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws2.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
        for col in ws2.columns:
            letter = get_column_letter(col[0].column)
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws2.column_dimensions[letter].width = min(max_len + 5, 60)
        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = ws2.dimensions

    return filename


@app.route("/api/export_xlsx")
def api_export_xlsx():
    if request.args.get("refresh") == "1":
        clear_cache()
    from_date = request.args.get("from") or ""
    try:
        if from_date:
            datetime.strptime(from_date, "%Y-%m-%d")
        else:
            today = datetime.now()
            from_date = f"{today.year:04d}-{today.month:02d}-06"
    except ValueError:
        today = datetime.now()
        from_date = f"{today.year:04d}-{today.month:02d}-06"
    report = build_from_date_report(from_date)
    filename = write_from_date_excel(report)
    return send_file(os.path.abspath(filename), as_attachment=True)


# =========================================================
# RUN
# =========================================================
if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5050)
