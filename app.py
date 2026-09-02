from flask import Flask, jsonify, render_template, send_file, request
import xmlrpc.client
import pandas as pd
from PIL import Image
from io import BytesIO
import os
import re
import base64
import uuid
import time
from dotenv import load_dotenv
from datetime import datetime, timedelta
import pdfkit
from collections import Counter, defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

ODOO_URL = os.getenv("ODOO_URL")
ODOO_DB = os.getenv("ODOO_DB")
ODOO_USERNAME = os.getenv("ODOO_USERNAME")
ODOO_PASSWORD = os.getenv("ODOO_API_KEY") or os.getenv("ODOO_PASSWORD")

app = Flask(__name__)

common = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/common", allow_none=True)
uid = common.authenticate(ODOO_DB, ODOO_USERNAME, ODOO_PASSWORD, {})
if not uid:
    raise SystemExit("AUTH FAILED - check .env credentials")

models = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/object", allow_none=True)

# =========================================================
# CACHING
# =========================================================

_FIELDS_CACHE = {}
_MASTER_CACHE = {"ts": 0.0, "data": None}
_SALES_CACHE = {"ts": 0.0, "data": None}
MASTER_CACHE_TTL = 180.0


def clear_cache():
    _FIELDS_CACHE.clear()
    _MASTER_CACHE["ts"] = 0.0
    _MASTER_CACHE["data"] = None
    _SALES_CACHE["ts"] = 0.0
    _SALES_CACHE["data"] = None


# =========================================================
# ODOO 19 SCHEMA KNOWLEDGE (verified against live DB)
# =========================================================
# product.template.type selection: consu=Goods, service=Service, combo=Combo
#  -> 392 Goods (357 storable), 13 Service, 0 Combo
#
# SELLING PRICE  -> product.product.lst_price (STORED, variant-level).
#                   product.template.list_price is COMPUTED (first variant).
#                   price_extra is ALREADY folded into lst_price - never add it.
#                   A lst_price of 1.0 is the Odoo placeholder -> treat as missing.
# BUYING PRICE   -> product.product.standard_price (STORED, variant-level).
#                   template.standard_price is computed (first variant).
#                   product.supplierinfo (13 rows, USD) is NOT a usable KES cost.
# STOCK          -> product.product.qty_available (on-hand). Template qty == sum
#                   of its variants (verified, 0 mismatches). Services have no
#                   quants and qty=0 -> stock checks do NOT apply to them.
# DESCRIPTIONS   -> goods: description_ecommerce (81% filled). website_description,
#                   description_sale, public_description are EMPTY in this DB.
# IMAGE          -> product.template.image_512 / image_1920.
# PUBLISHED      -> product.template.website_published.
# SALES          -> POS is the dominant channel (619 orders / 1001 lines; 616 of
#                   651 pickings). sale.order 181 (32 confirmed). Line fields:
#                   pos.order.line.price_subtotal (NO price_total in this build),
#                   sale.order.line.price_total + product_uom_qty. Only paid/done
#                   (POS) and sale (SO) states count toward revenue.
# =========================================================

TYPE_LABELS = {"consu": "Goods", "service": "Service", "combo": "Combo"}

GOODS_TYPES = ("consu", "service", "combo")


def product_type_label(raw_type):
    return TYPE_LABELS.get(raw_type, str(raw_type or "").title() or "Unknown")


# =========================================================
# ODOO HELPERS
# =========================================================

def safe_call(model, method, args=None, kwargs=None, retries=3):
    """Call Odoo with automatic retries so transient connection failures
    cannot silently wipe out an entire dataset."""
    if args is None:
        args = []
    if kwargs is None:
        kwargs = {}
    last_err = None
    for attempt in range(retries):
        try:
            return models.execute_kw(ODOO_DB, uid, ODOO_PASSWORD, model, method, args, kwargs)
        except Exception as e:
            last_err = e
            print(f"\nODOO ERROR | model={model} method={method} attempt={attempt + 1}/{retries} | {e}")
            time.sleep(1 + attempt * 2)
    print(f"\nODOO FAILED PERMANENTLY | model={model} method={method} | {last_err}")
    return []


def get_valid_model_fields(model, preferred_fields):
    """Return only the requested fields that actually exist on the Odoo model."""
    if not preferred_fields:
        return []
    if model not in _FIELDS_CACHE:
        try:
            _FIELDS_CACHE[model] = models.execute_kw(
                ODOO_DB, uid, ODOO_PASSWORD, model, "fields_get",
                [], {"attributes": ["name"]},
            )
        except Exception:
            _FIELDS_CACHE[model] = None
    field_map = _FIELDS_CACHE.get(model)
    if not field_map:
        return preferred_fields
    available = set(field_map.keys())
    valid = [f for f in preferred_fields if f in available]
    return valid or preferred_fields


def search_read(model, fields, domain=None, limit=100000):
    return safe_call(model, "search_read", [domain or []], {"fields": fields, "limit": limit})


# =========================================================
# TEXT / IMAGE HELPERS
# =========================================================

def clean(text):
    if not text:
        return ""
    text = re.sub(r"<[^>]+>", " ", str(text))
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def image_src(img):
    if not img:
        return ""
    if isinstance(img, str) and img.startswith("data:image"):
        return img
    if isinstance(img, bytes):
        img = base64.b64encode(img).decode("utf-8")
    if img.startswith("UklGR"):
        return f"data:image/webp;base64,{img}"
    if img.startswith("/9j/"):
        return f"data:image/jpeg;base64,{img}"
    if img.startswith("iVBOR"):
        return f"data:image/png;base64,{img}"
    if img.startswith("R0lGOD"):
        return f"data:image/gif;base64,{img}"
    return f"data:image/png;base64,{img}"


def get_image_raw(rec):
    if not rec:
        return None
    for f in ("image_1920", "image_1024", "image_512", "image_256", "image_128", "image_small", "image_medium"):
        if rec.get(f):
            return rec.get(f)
    return None


# Odoo 19 fact: goods carry their sales copy in description_ecommerce;
# the other description fields are empty in this database.
DESC_FIELDS = [
    "description_ecommerce",
    "website_description",
    "description_sale",
    "public_description",
    "description",
]


def find_best_description(product):
    best_text, best_source = "", ""
    for field in DESC_FIELDS:
        value = product.get(field)
        if not value:
            continue
        cleaned = clean(value)
        if len(cleaned) < 5:
            continue
        if len(cleaned) > len(best_text):
            best_text, best_source = cleaned, field
    return best_text, best_source


# =========================================================
# PRICE INTELLIGENCE
# =========================================================

# Junk placeholder price Odoo writes when no sale price is set.
JUNK_PRICE = 1.0


def _to_float(value, min_val=1.0):
    try:
        f = float(value)
        return f if f > min_val else None
    except (TypeError, ValueError):
        return None


def get_sale_price_from_record(rec, product_type="consu"):
    """Stored selling price. For goods the stored field is lst_price
    (template list_price is computed from the first variant).
    price_extra is already included in lst_price - never add it.
    A 1.0 placeholder is NOT a real price."""
    if not rec:
        return None
    if product_type == "service":
        for key in ("lst_price", "list_price"):
            val = _to_float(rec.get(key))
            if val:
                return val
        return None
    for key in ("lst_price", "list_price", "sale_price", "public_price", "price_unit", "regular_price"):
        if key in rec:
            val = _to_float(rec.get(key))
            if val:
                return val
    for key, val in rec.items():
        if not isinstance(key, str):
            continue
        lower = key.lower()
        if not any(w in lower for w in ("price", "sale", "list", "public", "regular")):
            continue
        if any(w in lower for w in ("cost", "purchase", "supplier", "standard")):
            continue
        valf = _to_float(val)
        if valf:
            return valf
    return None


def get_cost_from_record(rec, product_type="consu"):
    """Stored buying price. standard_price on product.product is the stored
    cost; supplierinfo prices are USD and are never used directly."""
    if not rec:
        return None
    for key in ("standard_price", "cost", "average_cost", "avg_cost", "price_cost", "supplier_price"):
        if key in rec:
            val = _to_float(rec.get(key))
            if val:
                return val
    if product_type == "service":
        return None
    for key, val in rec.items():
        if not isinstance(key, str):
            continue
        lower = key.lower()
        if "price" not in lower and "cost" not in lower:
            continue
        if any(w in lower for w in ("sale", "list", "public", "retail", "unit")):
            continue
        valf = _to_float(val)
        if valf:
            return valf
    return None


# =========================================================
# DEFECT ENGINE
# =========================================================

def defects(rec, product_type, storable, description, sale, cost, stock, image, published):
    """Type-aware defect engine (raw Odoo types: consu / service / combo).
    - Goods (consu): stock defects apply to storable items.
    - Services: never flagged for stock, buy price, image or margin - they are
      backoffice line items (expense codes), not purchased inventory.
    - Combos: no stock of their own."""
    d = []
    if product_type == "consu":
        if storable:
            if stock is None or stock <= 0:
                d.append("Out Of Stock")
            elif stock <= 5:
                d.append("Low Stock")
        if not description:
            d.append("Missing Description")
        elif len(description.strip()) < 30:
            d.append("Weak Description")
        if not image:
            d.append("Missing Image")
    elif product_type == "service":
        # services have no stock, no cost, no inventory image rules
        pass
    elif product_type == "combo":
        if not description:
            d.append("Missing Description")
    if sale is None or sale <= 0:
        d.append("Missing Selling Price")
    if product_type != "service":
        if cost is None or cost <= 0:
            d.append("Missing Buying Price")
        if sale and cost and sale > 0 and cost > 0:
            margin = ((sale - cost) / sale) * 100
            if margin < 10:
                d.append("Low Margin")
    if not published:
        d.append("Unpublished")
    return d


# =========================================================
# MASTER DATA
# =========================================================

def fetch_master_data():
    now = time.time()
    if _MASTER_CACHE["data"] and now - _MASTER_CACHE["ts"] < MASTER_CACHE_TTL:
        return _MASTER_CACHE["data"]

    template_fields = get_valid_model_fields("product.template", [
        "id", "name", "default_code", "barcode", "categ_id", "type", "is_storable",
        "list_price", "standard_price", "qty_available", "image_512", "create_date",
        "website_published", "description", "description_sale",
        "website_description", "public_description", "description_ecommerce",
    ])
    variant_fields = get_valid_model_fields("product.product", [
        "id", "name", "display_name", "product_tmpl_id", "default_code",
        "barcode", "lst_price", "list_price", "standard_price", "price_extra",
        "qty_available", "image_512", "create_date",
    ])

    templates = search_read("product.template", template_fields)
    variants = search_read("product.product", variant_fields)

    # Internal-usage locations only (excludes supplier / transit / adjustment).
    locations = search_read("stock.location", ["id", "name", "complete_name", "usage"])
    loc_display = {l["id"]: (l.get("complete_name") or l.get("name") or "") for l in locations}
    internal_ids = {l["id"] for l in locations if l.get("usage") == "internal"}
    if not internal_ids:
        internal_ids = None

    quants = search_read("stock.quant", ["product_id", "location_id", "quantity"])

    location_map = {}
    for q in quants:
        if not q.get("product_id") or not q.get("location_id"):
            continue
        loc_id = q["location_id"][0]
        if internal_ids is not None and loc_id not in internal_ids:
            continue
        pid = q["product_id"][0]
        location_map.setdefault(pid, []).append({
            "id": loc_id,
            "name": loc_display.get(loc_id) or q["location_id"][1],
            "qty": q.get("quantity", 0),
        })

    result = {
        "templates": templates,
        "variants": variants,
        "location_map": location_map,
        "fetch_health": {
            "templates": len(templates),
            "variants": len(variants),
            "quants": len(quants),
            "locations": len(location_map),
        },
    }
    _MASTER_CACHE["ts"] = now
    _MASTER_CACHE["data"] = result
    return result


# =========================================================
# SALES DATA (POS + web orders)
# =========================================================

POS_OK_STATES = ("paid", "done")
SO_OK_STATES = ("sale",)

_EMPTY_SALES = {"units": 0.0, "revenue": 0.0, "units30": 0.0, "rev30": 0.0, "last": ""}


def fetch_sales_data():
    """Per-variant sales from pos.order.line + sale.order.line plus a
    6-month revenue trend. POS is the dominant channel in this DB."""
    now = time.time()
    if _SALES_CACHE["data"] and now - _SALES_CACHE["ts"] < MASTER_CACHE_TTL:
        return _SALES_CACHE["data"]

    pos_line_fields = get_valid_model_fields("pos.order.line", ["id", "product_id", "qty", "price_subtotal", "order_id"])
    pos_order_fields = get_valid_model_fields("pos.order", ["id", "date_order", "state"])
    so_line_fields = get_valid_model_fields("sale.order.line", ["id", "product_id", "product_uom_qty", "price_total", "order_id"])
    so_order_fields = get_valid_model_fields("sale.order", ["id", "date_order", "state"])

    pos_lines = search_read("pos.order.line", pos_line_fields)
    pos_orders = {o["id"]: o for o in search_read("pos.order", pos_order_fields)}
    so_lines = search_read("sale.order.line", so_line_fields)
    so_orders = {o["id"]: o for o in search_read("sale.order", so_order_fields)}

    by_product = {}
    monthly = defaultdict(float)
    cut30 = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")

    def _fold(lines, orders, ok_states, qty_field, rev_field):
        for line in lines:
            pid = line.get("product_id")
            if not pid:
                continue
            pid = pid[0]
            order = None
            oid = line.get("order_id")
            if oid:
                order = orders.get(oid[0])
            if not order or order.get("state") not in ok_states:
                continue
            s = by_product.setdefault(pid, dict(_EMPTY_SALES))
            qty = float(line.get(qty_field) or 0)
            rev = float(line.get(rev_field) or 0)
            s["units"] += qty
            s["revenue"] += rev
            date_str = str(order.get("date_order") or "")
            if date_str[:10] >= cut30:
                s["units30"] += qty
                s["rev30"] += rev
            if date_str > s["last"]:
                s["last"] = date_str
            if len(date_str) >= 7:
                monthly[date_str[:7]] += rev

    _fold(pos_lines, pos_orders, POS_OK_STATES, "qty", "price_subtotal")
    _fold(so_lines, so_orders, SO_OK_STATES, "product_uom_qty", "price_total")

    today = datetime.now()
    y, m = today.year, today.month
    months = []
    for _ in range(6):
        months.append(f"{y:04d}-{m:02d}")
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    months.reverse()
    trend = [{"month": k, "revenue": round(monthly.get(k, 0.0), 2)} for k in months]

    result = {"by_product": by_product, "trend": trend, "lines": len(pos_lines) + len(so_lines)}
    _SALES_CACHE["ts"] = now
    _SALES_CACHE["data"] = result
    return result


def sales_for(pid):
    s = _SALES_CACHE["data"]["by_product"].get(pid) if _SALES_CACHE["data"] else None
    if not s:
        return dict(_EMPTY_SALES)
    return s


def aggregate_sales(linked_variants):
    agg = dict(_EMPTY_SALES)
    for vv in linked_variants:
        s = sales_for(vv["id"])
        for k in ("units", "revenue", "units30", "rev30"):
            agg[k] += s[k]
        if s["last"] > agg["last"]:
            agg["last"] = s["last"]
    return agg


# =========================================================
# PROCESS -> ROWS
# =========================================================

def _majority_value(values):
    """Return the unique value that holds >50% of the family, else None."""
    if not values:
        return None
    uniq = set(values)
    if len(uniq) == 1:
        return values[0]
    top, top_n = Counter(values).most_common(1)[0]
    return top if top_n > len(values) / 2 else None


def _to_num(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def build_rows(include_images=False):
    data = fetch_master_data()
    fetch_sales_data()
    templates, variants = data["templates"], data["variants"]
    location_map = data["location_map"]

    template_map = {t["id"]: t for t in templates}
    template_descs = {tid: find_best_description(t)[0] for tid, t in template_map.items()}
    template_type = {t["id"]: t.get("type", "consu") for t in templates}
    template_storable = {t["id"]: bool(t.get("is_storable")) for t in templates}

    variant_groups = {}
    for v in variants:
        tid = v.get("product_tmpl_id")[0] if v.get("product_tmpl_id") else None
        variant_groups.setdefault(tid, []).append(v)

    # Per-template family prices: a value is only inherited by a price-less
    # sibling when it is the family MAJORITY (>50%). Mixed families (e.g.
    # endmills priced by size) are never inherited - the variant is flagged
    # Missing instead of given a price that does not apply to it.
    family = {}
    for tid, linked in variant_groups.items():
        ptype = template_type.get(tid, "consu")
        sells = [s for s in (get_sale_price_from_record(vv, ptype) for vv in linked) if s]
        buys = [b for b in (get_cost_from_record(vv, ptype) for vv in linked) if b]
        family[tid] = {
            "sell_majority": _majority_value(sells),
            "buy_majority": _majority_value(buys),
        }

    rows = []

    for p in templates:
        tmpl_id = p["id"]
        ptype = template_type.get(tmpl_id, "consu")
        type_label = product_type_label(ptype)
        storable = template_storable.get(tmpl_id, False)
        linked = variant_groups.get(tmpl_id, [])
        fam = family.get(tmpl_id, {})
        is_goods = ptype == "consu"

        sell_prices = [s for s in (get_sale_price_from_record(vv, ptype) for vv in linked) if s]
        buy_prices = [b for b in (get_cost_from_record(vv, ptype) for vv in linked) if b]

        # Stock: qty_available is the authoritative on-hand figure.
        stock_total = sum(float(vv.get("qty_available") or 0) for vv in linked)
        if not linked:
            stock_total = float(p.get("qty_available") or 0)
        stock = stock_total if is_goods else None

        template_locations = {}
        for vv in linked:
            for loc in location_map.get(vv["id"], []):
                lid = loc["id"]
                if lid not in template_locations:
                    template_locations[lid] = {"id": lid, "name": loc["name"], "qty": 0}
                template_locations[lid]["qty"] += loc["qty"]
        template_locations = list(template_locations.values())
        location_summary = ", ".join(f'{l["name"]} ({l["qty"]})' for l in template_locations)

        if sell_prices:
            lo, hi = min(sell_prices), max(sell_prices)
            price_display = lo if lo == hi else f"{lo} → {hi}"
        else:
            price_display = get_sale_price_from_record(p, ptype)

        desc, source = find_best_description(p)
        published = bool(p.get("website_published"))

        sell_for_defect = min([x for x in sell_prices if x and x > 0]) if sell_prices else get_sale_price_from_record(p, ptype)

        # Buying price: only a real, unambiguous value counts.
        cost_for_defect = None
        cost_source = ""
        if is_goods:
            if fam.get("buy_majority"):
                cost_for_defect = fam["buy_majority"]
                cost_source = "family standard_price"
            elif buy_prices:
                cost_for_defect = min(buy_prices)
                cost_source = "variant standard_price"
            elif not buy_prices:
                tbp = get_cost_from_record(p, ptype)
                if tbp:
                    cost_for_defect = tbp
                    cost_source = "standard_price"

        d = defects(p, ptype, storable, desc, sell_for_defect, cost_for_defect, stock, get_image_raw(p), published)
        if len(set(sell_prices)) > 3:
            d.append("Variant Price Inconsistency")

        sell_source = ""
        if sell_prices:
            sell_source = "lst_price" if any(vv.get("lst_price") and vv.get("lst_price") > 0 for vv in linked) else "variant list_price"

        tsales = aggregate_sales(linked)
        stock_value = round(stock * cost_for_defect, 2) if is_goods and stock is not None and cost_for_defect else None
        days_of_stock = round(stock / (tsales["units30"] / 30), 1) if is_goods and stock is not None and stock > 0 and tsales["units30"] > 0 else None

        rows.append({
            "#": len(rows) + 1,
            "id": tmpl_id,
            "name": p.get("name"),
            "type": "TEMPLATE",
            "product_type": type_label,
            "raw_type": ptype,
            "is_storable": storable,
            "category": p["categ_id"][1] if p.get("categ_id") else "",
            "buy": cost_for_defect,
            "sell": price_display,
            "sell_raw": sell_for_defect,
            "buy_raw": cost_for_defect,
            "margin": round(((sell_for_defect - cost_for_defect) / sell_for_defect) * 100, 1)
            if sell_for_defect and cost_for_defect and sell_for_defect > 0 and is_goods else None,
            "price_source": (sell_source + " / " + cost_source) if sell_source else cost_source,
            "stock": stock,
            "stock_na": stock is None,
            "stock_value": stock_value,
            "days_of_stock": days_of_stock,
            "locations": template_locations,
            "location_summary": location_summary,
            "published": "YES" if published else "NO",
            "image": "YES" if get_image_raw(p) else "NO",
            "image_raw": get_image_raw(p) if include_images else "",
            "variant_count": len(linked),
            "variant_of": "",
            "description": desc,
            "description_source": source,
            "create_date": p.get("create_date"),
            "units_30d": round(tsales["units30"], 2),
            "revenue_30d": round(tsales["rev30"], 2),
            "units_total": round(tsales["units"], 2),
            "revenue_total": round(tsales["revenue"], 2),
            "avg_sale_price": round(tsales["revenue"] / tsales["units"], 2) if tsales["units"] else None,
            "last_sale_date": tsales["last"] or None,
            "abc_class": "No Sales",
            "status": "GOOD" if len(d) == 0 else ", ".join(d),
            "defects": d,
            "template_id": tmpl_id,
        })

    for v in variants:
        tmpl_id_v = v.get("product_tmpl_id")[0] if v.get("product_tmpl_id") else None
        tmpl = template_map.get(tmpl_id_v, {})
        template_name = tmpl.get("name", "")
        tmpl_desc = template_descs.get(tmpl_id_v, "")
        fam = family.get(tmpl_id_v, {})
        ptype = template_type.get(tmpl_id_v, "consu")
        type_label = product_type_label(ptype)
        storable = template_storable.get(tmpl_id_v, False)
        is_goods = ptype == "consu"

        # SELLING PRICE: own stored value only. A mixed family is never inherited.
        sell_price, sell_source = None, ""
        sp = get_sale_price_from_record(v, ptype)
        if sp:
            sell_price, sell_source = sp, "lst_price" if v.get("lst_price") else "list_price"
        else:
            tsp = get_sale_price_from_record(tmpl, ptype)
            if tsp:
                sell_price, sell_source = tsp, "template list_price"
            elif fam.get("sell_majority"):
                sell_price, sell_source = fam["sell_majority"], "family lst_price"

        # BUYING PRICE: own stored value only. A mixed family is never inherited.
        buy_price, buy_source = None, ""
        if is_goods:
            bp = get_cost_from_record(v, ptype)
            if bp:
                buy_price, buy_source = bp, "standard_price"
            else:
                tbp = get_cost_from_record(tmpl, ptype)
                if tbp:
                    buy_price, buy_source = tbp, "template standard_price"
                elif fam.get("buy_majority"):
                    buy_price, buy_source = fam["buy_majority"], "family standard_price"

        # Stock: qty_available is the authoritative on-hand figure. Services
        # have no stock (qty is always 0 and they hold no quants).
        stock = None
        if is_goods:
            stock = float(v.get("qty_available") or 0)
            if stock == 0:
                stock = sum(l["qty"] for l in location_map.get(v["id"], []))
        stock_display = stock if is_goods else None

        locations = location_map.get(v["id"], [])
        location_text = ", ".join(f'{l["name"]} ({l["qty"]})' for l in locations)

        variant_has_image = bool(get_image_raw(v)) or bool(get_image_raw(tmpl))

        d = defects(v, ptype, storable, tmpl_desc, sell_price, buy_price, stock, variant_has_image, published=bool(tmpl.get("website_published")))

        s = sales_for(v["id"])
        stock_value = round(stock * buy_price, 2) if is_goods and stock is not None and buy_price else None
        days_of_stock = round(stock / (s["units30"] / 30), 1) if is_goods and stock is not None and stock > 0 and s["units30"] > 0 else None

        rows.append({
            "#": len(rows) + 1,
            "id": v.get("id"),
            "name": v.get("display_name") or v.get("name"),
            "type": "VARIANT",
            "product_type": type_label,
            "raw_type": ptype,
            "is_storable": storable,
            "category": tmpl.get("categ_id")[1] if tmpl.get("categ_id") else "",
            "buy": buy_price,
            "sell": sell_price,
            "sell_raw": sell_price,
            "buy_raw": buy_price,
            "margin": round(((sell_price - buy_price) / sell_price) * 100, 1)
            if sell_price and buy_price and sell_price > 0 and is_goods else None,
            "price_source": " / ".join(x for x in (sell_source, buy_source) if x),
            "stock": stock_display,
            "stock_na": stock_display is None,
            "stock_value": stock_value,
            "days_of_stock": days_of_stock,
            "locations": locations,
            "location_summary": location_text,
            "published": "YES" if tmpl.get("website_published") else "NO",
            "image": "YES" if variant_has_image else "NO",
            "image_raw": (get_image_raw(v) or get_image_raw(tmpl)) if include_images else "",
            "variant_count": 0,
            "variant_of": template_name,
            "description": tmpl_desc,
            "description_source": "",
            "create_date": v.get("create_date"),
            "units_30d": round(s["units30"], 2),
            "revenue_30d": round(s["rev30"], 2),
            "units_total": round(s["units"], 2),
            "revenue_total": round(s["revenue"], 2),
            "avg_sale_price": round(s["revenue"] / s["units"], 2) if s["units"] else None,
            "last_sale_date": s["last"] or None,
            "abc_class": "No Sales",
            "status": "GOOD" if len(d) == 0 else ", ".join(d),
            "defects": d,
            "template_id": tmpl_id_v,
        })

    # ABC classification by last-30d revenue (cumulative share).
    # A = first 70% of revenue, B = next 25%, C = rest. No sales -> "No Sales".
    ranked = sorted([r for r in rows if (r.get("revenue_30d") or 0) > 0], key=lambda r: -r["revenue_30d"])
    total = sum(r["revenue_30d"] for r in ranked)
    cum = 0.0
    for r in ranked:
        cum += r["revenue_30d"]
        share = cum / total if total else 0
        r["abc_class"] = "A" if share <= 0.70 else ("B" if share <= 0.95 else "C")

    return rows


def apply_filters(rows):
    issue = request.args.get("issue")
    search = request.args.get("search")
    location = request.args.get("location")
    category = request.args.get("category")
    type_filter = request.args.get("type")
    ptype_filter = request.args.get("ptype")
    status = request.args.get("status")
    min_margin = request.args.get("min_margin")
    abc = request.args.get("abc")
    has_sales = request.args.get("has_sales")
    min_stock = request.args.get("min_stock")
    max_stock = request.args.get("max_stock")

    if issue:
        rows = [r for r in rows if issue in r.get("defects", [])]
    if search:
        s = search.lower()
        rows = [r for r in rows if s in (r.get("name") or "").lower()
                or s in (r.get("variant_of") or "").lower()
                or s in (r.get("category") or "").lower()]
    if location:
        rows = [r for r in rows if any(str(loc.get("id")) == str(location) for loc in r.get("locations", []))]
    if category:
        rows = [r for r in rows if r.get("category") == category]
    if type_filter:
        rows = [r for r in rows if r.get("type") == type_filter.upper()]
    if ptype_filter:
        rows = [r for r in rows if r.get("product_type") == ptype_filter]
    if status:
        if status.upper() == "GOOD":
            rows = [r for r in rows if r.get("status") == "GOOD"]
        elif status.upper() == "ISSUES":
            rows = [r for r in rows if r.get("status") != "GOOD"]
    if min_margin:
        try:
            mm = float(min_margin)
            rows = [r for r in rows if r.get("margin") is not None and r["margin"] < mm]
        except ValueError:
            pass
    if abc:
        rows = [r for r in rows if (r.get("abc_class") or "") == abc]
    if has_sales:
        rows = [r for r in rows if (r.get("revenue_total") or 0) > 0 or (r.get("units_total") or 0) > 0]
    if min_stock:
        try:
            ms = float(min_stock)
            rows = [r for r in rows if r.get("stock") is not None and r["stock"] >= ms]
        except ValueError:
            pass
    if max_stock:
        try:
            mx = float(max_stock)
            rows = [r for r in rows if r.get("stock") is not None and r["stock"] <= mx]
        except ValueError:
            pass
    return rows


def build_export_frame(rows, meta):
    df = pd.DataFrame([{
        "#": r["#"],
        "Type": r["type"],
        "Product Type": r.get("product_type", ""),
        "Storable": "YES" if r.get("is_storable") else "NO",
        "Name": r["name"],
        "Category": r.get("category") or r.get("variant_of") or "",
        "Variant Of": r.get("variant_of", ""),
        "Buy Price": r.get("buy_raw"),
        "Sell Price": r.get("sell_raw"),
        "Margin %": r.get("margin"),
        "Price Source": r.get("price_source", ""),
        "Stock": r.get("stock") if r.get("stock") is not None else "N/A",
        "Stock Value": r.get("stock_value"),
        "Units 30d": r.get("units_30d"),
        "Revenue 30d": r.get("revenue_30d"),
        "Units Total": r.get("units_total"),
        "Revenue Total": r.get("revenue_total"),
        "ABC": r.get("abc_class", ""),
        "Locations": r.get("location_summary", ""),
        "Published": r.get("published", ""),
        "Image": r.get("image", ""),
        "Status": r.get("status", ""),
        "Defects": ", ".join(r.get("defects", [])),
    } for r in rows])
    return df


def write_excel(df, rows, meta):
    os.makedirs("exports", exist_ok=True)
    filename = f"exports/products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    defect_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    good_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        meta_df = pd.DataFrame(list(meta.items()), columns=["key", "value"])
        meta_df.to_excel(writer, index=False, sheet_name="Filters")
        ws_meta = writer.sheets["Filters"]
        ws_meta.column_dimensions["A"].width = 22
        ws_meta.column_dimensions["B"].width = 60
        for cell in ws_meta[1]:
            cell.font = Font(bold=True)

        df.to_excel(writer, index=False, sheet_name="Products")
        ws = writer.sheets["Products"]
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows(min_row=2):
            status = str(row[ws.max_column - 2].value or "")
            fill = good_fill if status == "GOOD" else defect_fill
            for cell in row:
                cell.border = border
                if status != "GOOD":
                    cell.fill = fill
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max_len + 5, 80)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    return filename


# =========================================================
# ROUTES
# =========================================================

@app.route("/")
def home():
    return render_template("dashboard.html")


@app.route("/api/products")
def api_products():
    if request.args.get("refresh") == "1":
        clear_cache()
    rows = build_rows(include_images=request.args.get("images") == "1")
    rows = apply_filters(rows)
    return jsonify(rows)


@app.route("/api/product_image")
def api_product_image():
    tid = request.args.get("id", type=int)
    if not tid:
        return jsonify({"image": ""})
    t = safe_call("product.template", "read", [[tid]], {"fields": ["image_512", "image_1920"]})
    if t and t[0]:
        img = t[0].get("image_512") or t[0].get("image_1920")
        if img:
            return jsonify({"image": image_src(img)})
    return jsonify({"image": ""})


@app.route("/api/analytics")
def api_analytics():
    if request.args.get("refresh") == "1":
        clear_cache()
    rows = build_rows()
    data = fetch_master_data()

    def f(x):
        return x or 0

    template_variant = Counter(r["type"] for r in rows)
    type_dist = Counter(r["product_type"] for r in rows)
    goods = [r for r in rows if r["product_type"] == "Goods"]
    stock_value_total = sum(f(r.get("stock_value")) for r in goods)
    revenue_30 = sum(f(r.get("revenue_30d")) for r in rows)
    units_30 = sum(f(r.get("units_30d")) for r in rows)

    defects_counter = Counter()
    for r in rows:
        for d in r.get("defects", []):
            defects_counter[d] += 1

    cat_map = {}
    for r in rows:
        c = r.get("category") or "Uncategorized"
        if c not in cat_map:
            cat_map[c] = {
                "category": c, "templates": 0, "variants": 0, "products": 0,
                "stock_value": 0.0, "revenue_30d": 0.0, "units_30d": 0.0,
                "out_of_stock": 0, "low_stock": 0, "missing_price": 0,
                "margin_sum": 0.0, "margin_count": 0, "with_sell": 0, "with_stock": 0,
            }
        e = cat_map[c]
        e["products"] += 1
        e["templates" if r["type"] == "TEMPLATE" else "variants"] += 1
        e["stock_value"] += f(r.get("stock_value"))
        e["revenue_30d"] += f(r.get("revenue_30d"))
        e["units_30d"] += f(r.get("units_30d"))
        if "Out Of Stock" in r.get("defects", []):
            e["out_of_stock"] += 1
        if "Low Stock" in r.get("defects", []):
            e["low_stock"] += 1
        if "Missing Selling Price" in r.get("defects", []):
            e["missing_price"] += 1
        if r.get("margin") is not None:
            e["margin_sum"] += r["margin"]
            e["margin_count"] += 1
        if r.get("sell_raw"):
            e["with_sell"] += 1
        if r.get("stock") is not None:
            e["with_stock"] += 1
    by_category = []
    for e in cat_map.values():
        e["avg_margin"] = round(e["margin_sum"] / e["margin_count"], 1) if e["margin_count"] else None
        e.pop("margin_sum", None)
        e.pop("margin_count", None)
        e["stock_value"] = round(e["stock_value"], 2)
        e["revenue_30d"] = round(e["revenue_30d"], 2)
        by_category.append(e)
    by_category.sort(key=lambda e: -e["products"])

    price_buckets = [
        ("0 - 500", lambda v: v <= 500),
        ("500 - 1K", lambda v: 500 < v <= 1000),
        ("1K - 2.5K", lambda v: 1000 < v <= 2500),
        ("2.5K - 5K", lambda v: 2500 < v <= 5000),
        ("5K - 10K", lambda v: 5000 < v <= 10000),
        ("10K - 25K", lambda v: 10000 < v <= 25000),
        ("25K+", lambda v: v > 25000),
    ]
    bucket_counts = {k: 0 for k, _ in price_buckets}
    bucket_counts["Missing"] = 0
    for r in rows:
        v = r.get("sell_raw")
        placed = False
        if v:
            for label, test in price_buckets:
                if test(float(v)):
                    bucket_counts[label] += 1
                    placed = True
                    break
        if not placed:
            bucket_counts["Missing"] += 1
    price_dist = [{"bucket": k, "count": v} for k, v in bucket_counts.items()]

    top_sellers = sorted(
        [r for r in rows if r.get("revenue_30d")],
        key=lambda r: -r["revenue_30d"],
    )[:15]
    top = [{
        "name": r["name"], "revenue": r["revenue_30d"], "units": r["units_30d"],
        "category": r["category"], "sell": r["sell_raw"], "abc_class": r["abc_class"],
    } for r in top_sellers]

    abc_counts = Counter(r.get("abc_class") for r in rows)

    loc_qty = defaultdict(float)
    for pid, locs in data["location_map"].items():
        for l in locs:
            loc_qty[l["name"]] += l["qty"]
    locations = [{"name": k, "qty": round(v, 2)} for k, v in sorted(loc_qty.items(), key=lambda kv: -kv[1])]

    margins = [r["margin"] for r in rows if r.get("margin") is not None]
    kpis = {
        "products": len(rows),
        "templates": template_variant.get("TEMPLATE", 0),
        "variants": template_variant.get("VARIANT", 0),
        "goods": type_dist.get("Goods", 0),
        "services": type_dist.get("Service", 0),
        "storable": sum(1 for r in rows if r.get("is_storable")),
        "stock_value": round(stock_value_total, 2),
        "revenue_30d": round(revenue_30, 2),
        "units_30d": round(units_30, 2),
        "avg_margin": round(sum(margins) / len(margins), 1) if margins else None,
        "published": sum(1 for r in rows if r.get("published") == "YES"),
        "good_rows": sum(1 for r in rows if r.get("status") == "GOOD"),
        "issue_rows": sum(1 for r in rows if r.get("status") != "GOOD"),
        "out_of_stock": defects_counter.get("Out Of Stock", 0),
        "low_stock": defects_counter.get("Low Stock", 0),
        "missing_price": defects_counter.get("Missing Selling Price", 0),
        "missing_buy": defects_counter.get("Missing Buying Price", 0),
        "missing_image": defects_counter.get("Missing Image", 0),
        "missing_desc": defects_counter.get("Missing Description", 0),
        "weak_desc": defects_counter.get("Weak Description", 0),
        "low_margin": defects_counter.get("Low Margin", 0),
        "unpublished": defects_counter.get("Unpublished", 0),
        "abc": dict(abc_counts),
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }

    return jsonify({
        "kpis": kpis,
        "by_category": by_category,
        "defects": [{"name": k, "count": v} for k, v in defects_counter.most_common()],
        "type_dist": dict(type_dist),
        "price_dist": price_dist,
        "top_sellers": top,
        "sales_trend": fetch_sales_data()["trend"],
        "locations": locations,
        "fetch_health": data["fetch_health"],
    })


@app.route("/api/options")
def api_options():
    if request.args.get("refresh") == "1":
        clear_cache()
    rows = build_rows()
    categories = sorted({r["category"] for r in rows if r["category"]})
    locations = {}
    for r in rows:
        for loc in r.get("locations", []):
            locations[loc["id"]] = loc["name"]
    product_types = sorted({r["product_type"] for r in rows if r.get("product_type")})
    return jsonify({
        "categories": categories,
        "locations": [{"id": k, "name": v} for k, v in sorted(locations.items())],
        "product_types": product_types,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    })


@app.route("/export_report")
def export_report():
    rows = build_rows()
    rows = apply_filters(rows)
    df = build_export_frame(rows, {})
    meta = {
        "generated_at": datetime.now().isoformat(),
        "total_rows": len(rows),
        "issue_filter": request.args.get("issue") or "",
        "search_filter": request.args.get("search") or "",
        "location_filter": request.args.get("location") or "",
        "category_filter": request.args.get("category") or "",
        "type_filter": request.args.get("type") or "",
        "product_type_filter": request.args.get("ptype") or "",
        "status_filter": request.args.get("status") or "",
        "abc_filter": request.args.get("abc") or "",
        "min_margin_filter": request.args.get("min_margin") or "",
    }
    filename = write_excel(df, rows, meta)
    return send_file(os.path.abspath(filename), as_attachment=True)


# =========================================================
# CATALOGUE
# =========================================================

@app.route("/catalogue")
def catalogue_builder():
    return render_template("catalogue.html")


@app.route("/api/catalogue_products")
def catalogue_products():
    rows = build_rows(include_images=True)
    products = []
    for r in rows:
        # Services without a real sell price are internal expense codes
        # (Communication, Meals, Gifts...) - never catalogue fodder.
        if r.get("product_type") == "Service" and not (r.get("sell_raw") or 0):
            continue
        products.append({
            "id": r["id"] if r["type"] == "TEMPLATE" else f"variant_{r['id']}",
            "variant_id": r["id"] if r["type"] == "VARIANT" else None,
            "template_id": r.get("template_id"),
            "type": r["type"],
            "product_type": r.get("product_type", ""),
            "name": r.get("name", ""),
            "sell": r.get("sell_raw") or 0,
            "buy": r.get("buy_raw") or 0,
            "stock": r.get("stock") or 0,
            "category": r.get("category") or "",
            "description": r.get("description", ""),
            "description_source": r.get("description_source", ""),
            "image_1920": image_src(r.get("image_raw")),
            "published": r.get("published") == "YES",
            "defects": r.get("defects", []),
        })

    products.sort(key=lambda x: (x["name"] or "").lower())
    return jsonify({
        "total_products": len(products),
        "templates": len([x for x in products if x["type"] == "TEMPLATE"]),
        "variants": len([x for x in products if x["type"] == "VARIANT"]),
        "products": products,
    })


@app.route("/catalogue/review")
def catalogue_review():
    return render_template("review.html")


def catalogue_html(data, for_pdf=True):
    products = data.get("products", [])
    title = data.get("title", "Product Catalogue")
    website = data.get("website", "")
    contacts = data.get("contacts", "")
    company_description = data.get("companyDescription", "")
    logo = data.get("logo", "")
    qr = data.get("qr", "")

    grouped = {}
    for p in products:
        category = p.get("catalogue_category") or p.get("category") or "Uncategorized"
        grouped.setdefault(category, []).append(p)

    html = """<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
    html, body { margin: 0; padding: 20px; min-height: 100vh;
        font-family: 'Segoe UI', sans-serif; background: transparent; color: #f9fafb; }
    #pdf-backdrop { position: fixed; top: 0; left: 0; width: 100%; height: 100%;
        z-index: 0; background: linear-gradient(135deg, #0ea5e9 0%, #60a5fa 35%, #fb923c 75%, #f97316 100%); }
    .content-wrap { position: relative; z-index: 1; }
    .header { text-align: center; padding: 25px; margin-bottom: 25px;
        background: rgba(17, 24, 39, 0.75); border-radius: 16px;
        box-shadow: 0 4px 20px rgba(0,0,0,.25); }
    .logo { max-height: 180px; width: 360px; }
    .qr { max-height: 200px; width: 400px; }
    .title { font-size: 28px; font-weight: bold; color: #f9fafb; }
    .meta { color: #cbd5e1; margin-top: 4px; }
    .category { background: linear-gradient(90deg, #111827, #EA580C, #DC2626);
        color: white; padding: 12px 16px; margin-top: 30px; margin-bottom: 15px;
        font-size: 18px; font-weight: bold; border-radius: 10px; page-break-after: avoid; }
    .products-grid { width: 100%; font-size: 0; }
    .product-card { display: inline-block; width: 30%; margin: 1%; vertical-align: top;
        font-size: 12px; background: rgba(255,255,255,0.92);
        border: 1px solid rgba(255,255,255,0.3); border-radius: 14px; padding: 15px;
        box-sizing: border-box; box-shadow: 0 6px 18px rgba(0,0,0,.25);
        page-break-inside: avoid; }
    .img-small { width: 120px; height: 120px; object-fit: contain; display: block; margin: auto; }
    .img-medium { width: 180px; height: 180px; object-fit: contain; display: block; margin: auto; }
    .img-large { width: 240px; height: 240px; object-fit: contain; display: block; margin: auto; }
    .company-profile { margin: 25px 0; padding: 25px; background: rgba(17,24,39,0.75);
        border-left: 6px solid #EA580C; border-radius: 14px; line-height: 1.8;
        font-size: 13px; color: #e5e7eb; box-shadow: 0 4px 12px rgba(0,0,0,.25); }
    .company-profile h2 { margin: 0 0 15px 0; color: #ffffff; font-size: 22px; }
    .company-profile p { margin: 0; white-space: pre-wrap; text-align: justify; }
    .product-name { font-size: 16px; font-weight: 700; margin-top: 12px; color: #111827; }
    .old-price { text-decoration: line-through; color: #94a3b8; margin-top: 10px; }
    .new-price { color: #16a34a; font-size: 18px; font-weight: bold; }
    .discount { display: inline-block; margin-top: 8px; background: #DC2626; color: white;
        padding: 4px 10px; border-radius: 20px; font-size: 12px; }
    .description { margin-top: 20px; padding: 18px; background: rgba(255,255,255,0.9);
        border: 1px solid #e5e7eb; border-radius: 12px; line-height: 1.8; color: #374151;
        font-size: 13px; text-align: justify; white-space: pre-wrap; }
    .description h3 { margin-top: 0; margin-bottom: 12px; color: #111827; font-size: 18px; }
</style>
</head>
<body>
<div id="pdf-backdrop"></div>
<div class="content-wrap">
<div class="header">
    <div class="title">{{TITLE}}</div>
"""

    if logo:
        html += f'<img src="{logo}" class="logo"><br>'
    if website:
        html += f'<div class="meta">{website}</div>'
    if contacts:
        html += f'<div class="meta">{contacts}</div>'
    if qr:
        html += f'<br><img src="{qr}" class="qr">'
    html += "</div>"

    if company_description:
        html += f'<div class="company-profile"><h2>About Us</h2><p>{company_description}</p></div>'

    for category, items in grouped.items():
        html += f'<div class="category">{category}</div><div class="products-grid">'
        for p in items:
            sell = float(p.get("sell", 0) or 0)
            discount = float(p.get("discount", 0) or 0)
            old_price = sell + (sell * discount / 100)
            desc = p.get("description", "")
            if p.get("useSystemDesc") is False:
                desc = p.get("customDesc", "") or desc
            image = save_base64_image(p.get("image_1920", ""), for_pdf)
            image_class = {"small": "img-small", "medium": "img-medium", "large": "img-large"}.get(
                p.get("imageSize", "medium"), "img-medium")

            html += f'<div class="product-card">'
            if image:
                html += f'<img src="{image}" class="{image_class}">'
            if p.get("showName", True):
                html += f'<div class="product-name">{p.get("name", "")}</div>'
            if p.get("showPrice", True):
                if discount > 0:
                    html += f'<div class="old-price">KES {old_price:,.2f}</div>'
                html += f'<div class="new-price">KES {sell:,.2f}</div>'
            if p.get("showDiscount", True) and discount:
                html += f'<span class="discount">-{discount:.0f}%</span>'
            if p.get("showDescription", True) and desc:
                html += f'<div class="description">{desc}</div>'
            html += "</div>"
        html += "</div>"

    html += "</div></body></html>"
    return html.replace("{{TITLE}}", title)


def save_base64_image(base64_string, as_file=True):
    if not base64_string:
        return ""
    if not as_file:
        return base64_string if "," in base64_string else f"data:image/png;base64,{base64_string}"
    try:
        os.makedirs("temp_catalogue_images", exist_ok=True)
        if "," in base64_string:
            base64_string = base64_string.split(",")[1]
        image_bytes = base64.b64decode(base64_string)
        image = Image.open(BytesIO(image_bytes))
        filename = f"{uuid.uuid4()}.png"
        filepath = os.path.abspath(os.path.join("temp_catalogue_images", filename))
        image.save(filepath, format="PNG")
        return "file:///" + filepath.replace("\\", "/")
    except Exception as e:
        print("IMAGE CONVERSION ERROR:", e)
        return ""


@app.route("/catalogue/preview", methods=["POST"])
def catalogue_preview():
    data = request.json or {}
    return catalogue_html(data, for_pdf=False)


@app.route("/catalogue/download", methods=["POST"])
def catalogue_download():
    data = request.json or {}
    html = catalogue_html(data, for_pdf=True)

    wkhtmltopdf_path = os.getenv("WKHTMLTOPDF_PATH") or "/usr/bin/wkhtmltopdf"
    config = pdfkit.configuration(wkhtmltopdf=wkhtmltopdf_path)
    options = {
        "page-size": "A4",
        "encoding": "UTF-8",
        "background": "",
        "print-media-type": "",
        "enable-local-file-access": "",
        "images": "",
        "enable-javascript": "",
        "javascript-delay": "5000",
        "no-stop-slow-scripts": "",
        "margin-top": "5mm",
        "margin-bottom": "5mm",
        "margin-left": "5mm",
        "margin-right": "5mm",
    }
    pdf = pdfkit.from_string(html, False, configuration=config, options=options)
    os.makedirs("exports", exist_ok=True)
    filename = f"catalogue_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filepath = os.path.join("exports", filename)
    with open(filepath, "wb") as f:
        f.write(pdf)
    return send_file(os.path.abspath(filepath), as_attachment=True)


# =========================================================
# SALES REPORT
# =========================================================

_SALES_REPORT_CACHE = {"ts": 0.0, "data": None}
SALES_REPORT_CACHE_TTL = 300.0


def fetch_sales_lines():
    now = time.time()
    if _SALES_REPORT_CACHE["data"] and now - _SALES_REPORT_CACHE["ts"] < SALES_REPORT_CACHE_TTL:
        return _SALES_REPORT_CACHE["data"]

    pos_line_fields = get_valid_model_fields("pos.order.line", ["id", "product_id", "qty", "price_subtotal", "order_id"])
    pos_order_fields = get_valid_model_fields("pos.order", ["id", "date_order", "state"])
    so_line_fields = get_valid_model_fields("sale.order.line", ["id", "product_id", "product_uom_qty", "price_total", "order_id"])
    so_order_fields = get_valid_model_fields("sale.order", ["id", "date_order", "state"])
    variant_fields = get_valid_model_fields("product.product", ["id", "display_name", "name", "product_tmpl_id"])
    template_fields = get_valid_model_fields("product.template", ["id", "name", "categ_id"])

    pos_lines = search_read("pos.order.line", pos_line_fields)
    pos_orders = {o["id"]: o for o in search_read("pos.order", pos_order_fields)}
    so_lines = search_read("sale.order.line", so_line_fields)
    so_orders = {o["id"]: o for o in search_read("sale.order", so_order_fields)}
    variants = {v["id"]: v for v in search_read("product.product", variant_fields)}
    templates = {t["id"]: t for t in search_read("product.template", template_fields)}

    lines = []
    counts = {"pos": 0, "so": 0}

    def _fold(lines_raw, orders, ok_states, qty_field, rev_field, channel):
        for line in lines_raw:
            pid_ref = line.get("product_id")
            if not pid_ref:
                continue
            order = None
            oid = line.get("order_id")
            if oid:
                order = orders.get(oid[0])
            if not order or order.get("state") not in ok_states:
                continue
            date_str = str(order.get("date_order") or "")
            if not date_str:
                continue
            variant = variants.get(pid_ref[0]) or {}
            tid_ref = variant.get("product_tmpl_id")
            tid = tid_ref[0] if tid_ref else None
            template = templates.get(tid) or {}
            lines.append({
                "id": line.get("id"),
                "product_id": pid_ref[0],
                "template_id": tid,
                "product_name": variant.get("display_name") or variant.get("name") or "",
                "template_name": template.get("name") or "",
                "category": template.get("categ_id")[1] if template.get("categ_id") else "",
                "qty": float(line.get(qty_field) or 0),
                "revenue": float(line.get(rev_field) or 0),
                "date": date_str[:10],
                "month": date_str[:7],
                "channel": channel,
                "order_id": order["id"],
            })

    _fold(pos_lines, pos_orders, POS_OK_STATES, "qty", "price_subtotal", "POS")
    counts["pos"] = len(pos_lines)
    _fold(so_lines, so_orders, SO_OK_STATES, "product_uom_qty", "price_total", "Sales Order")
    counts["so"] = len(so_lines)

    lines.sort(key=lambda x: x["date"])
    months = []
    seen = set()
    for l in lines:
        if l["month"] not in seen:
            seen.add(l["month"])
            months.append(l["month"])

    result = {"lines": lines, "months": months, "counts": counts}
    _SALES_REPORT_CACHE["ts"] = now
    _SALES_REPORT_CACHE["data"] = result
    return result


def build_sales_payload(month=None):
    data = fetch_sales_lines()
    lines = data["lines"]
    months_sorted = data["months"]
    latest = months_sorted[-1] if months_sorted else None

    if month is None or (month not in months_sorted and month != "all"):
        month = latest or "all"

    monthly = defaultdict(lambda: {"revenue": 0.0, "units": 0.0, "orders": set()})
    for l in lines:
        m = monthly[l["month"]]
        m["revenue"] += l["revenue"]
        m["units"] += l["qty"]
        m["orders"].add((l["channel"], l["order_id"]))

    selected = lines if month == "all" else [l for l in lines if l["month"] == month]

    products = {}
    for l in lines:
        key = l["template_id"] or ("v%d" % l["product_id"])
        p = products.get(key)
        if p is None:
            p = {
                "key": key, "template_id": l["template_id"],
                "name": l["template_name"] or l["product_name"],
                "category": l["category"],
                "first_sale": l["date"], "last_sale": l["date"],
                "monthly": defaultdict(lambda: {"revenue": 0.0, "units": 0.0}),
                "channels": defaultdict(lambda: {"revenue": 0.0, "units": 0.0}),
                "variants": defaultdict(lambda: {"name": "", "revenue": 0.0, "units": 0.0}),
            }
            products[key] = p
        p["first_sale"] = min(p["first_sale"], l["date"])
        p["last_sale"] = max(p["last_sale"], l["date"])
        p["monthly"][l["month"]]["revenue"] += l["revenue"]
        p["monthly"][l["month"]]["units"] += l["qty"]
        p["channels"][l["channel"]]["revenue"] += l["revenue"]
        p["channels"][l["channel"]]["units"] += l["qty"]
        p["variants"][l["product_id"]]["name"] = l["product_name"]
        p["variants"][l["product_id"]]["revenue"] += l["revenue"]
        p["variants"][l["product_id"]]["units"] += l["qty"]

    sel_products = defaultdict(lambda: {"revenue": 0.0, "units": 0.0, "orders": set()})
    channel_totals = defaultdict(lambda: {"revenue": 0.0, "units": 0.0, "orders": set()})
    date_lines = defaultdict(lambda: defaultdict(lambda: {"name": "", "units": 0.0, "revenue": 0.0}))
    kpi_orders = set()
    for l in selected:
        key = l["template_id"] or ("v%d" % l["product_id"])
        sp = sel_products[key]
        sp["revenue"] += l["revenue"]
        sp["units"] += l["qty"]
        sp["orders"].add((l["channel"], l["order_id"]))
        ch = channel_totals[l["channel"]]
        ch["revenue"] += l["revenue"]
        ch["units"] += l["qty"]
        ch["orders"].add((l["channel"], l["order_id"]))
        kpi_orders.add((l["channel"], l["order_id"]))
        cell = date_lines[key][(l["date"], l["product_id"])]
        cell["name"] = l["product_name"]
        cell["units"] += l["qty"]
        cell["revenue"] += l["revenue"]

    product_rows = []
    for key, p in products.items():
        sp = sel_products.get(key)
        if not sp or not sp["orders"]:
            continue
        monthly_bd = [{"month": m, "revenue": round(d["revenue"], 2), "units": round(d["units"], 2)} for m, d in sorted(p["monthly"].items())]
        channel_bd = [{"name": n, "revenue": round(d["revenue"], 2), "units": round(d["units"], 2)} for n, d in sorted(p["channels"].items(), key=lambda kv: -kv[1]["revenue"])]
        variants = sorted([{"name": v["name"] or vid, "revenue": round(v["revenue"], 2), "units": round(v["units"], 2)} for vid, v in p["variants"].items()], key=lambda x: -x["revenue"])
        sales_by_date = sorted([{"date": d, "product": cell["name"], "units": round(cell["units"], 2), "revenue": round(cell["revenue"], 2)} for (d, _pid), cell in date_lines.get(key, {}).items()], key=lambda x: (x["date"], x["product"]))
        product_rows.append({
            "key": key, "template_id": p["template_id"], "name": p["name"],
            "category": p["category"] or "Uncategorized",
            "revenue": round(sp["revenue"], 2), "units": round(sp["units"], 2),
            "orders": len(sp["orders"]),
            "avg_price": round(sp["revenue"] / sp["units"], 2) if sp["units"] else None,
            "first_sale": p["first_sale"], "last_sale": p["last_sale"],
            "channel_breakdown": channel_bd, "monthly_breakdown": monthly_bd,
            "variants": variants, "sales_by_date": sales_by_date,
        })
    product_rows.sort(key=lambda x: -x["revenue"])

    trend = [{"month": m, "revenue": round(monthly[m]["revenue"], 2), "units": round(monthly[m]["units"], 2), "orders": len(monthly[m]["orders"])} for m in months_sorted]

    kpi_revenue = sum(sp["revenue"] for sp in sel_products.values())
    kpi_units = sum(sp["units"] for sp in sel_products.values())
    categories = {p["category"] for p in product_rows}

    return {
        "month": month, "months": months_sorted,
        "kpis": {
            "revenue": round(kpi_revenue, 2), "units": round(kpi_units, 2),
            "orders": len(kpi_orders), "products": len(product_rows),
            "lines": len(selected), "categories": len(categories),
            "avg_order_value": round(kpi_revenue / len(kpi_orders), 2) if kpi_orders else 0.0,
        },
        "channels": [{"name": n, "revenue": round(d["revenue"], 2), "units": round(d["units"], 2), "orders": len(d["orders"])} for n, d in sorted(channel_totals.items(), key=lambda kv: -kv[1]["revenue"])],
        "monthly_trend": trend,
        "products": product_rows,
        "fetch_health": {"lines": len(lines), "pos_lines": data["counts"]["pos"], "so_lines": data["counts"]["so"]},
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }


@app.route("/sales")
def sales_page():
    return render_template("sales/report.html")


@app.route("/api/sales")
def api_sales():
    if request.args.get("refresh") == "1":
        _FIELDS_CACHE.clear()
        _SALES_REPORT_CACHE["ts"] = 0.0
        _SALES_REPORT_CACHE["data"] = None
    return jsonify(build_sales_payload(request.args.get("month")))


# =========================================================
# RUN
# =========================================================
if __name__ == "__main__":
    port = int(os.getenv("PORT", 5000))
    debug = os.getenv("FLASK_DEBUG", "false").lower() == "true"
    app.run(debug=debug, host="0.0.0.0", port=port)