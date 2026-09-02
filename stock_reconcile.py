# =========================================================
# Stock reconciliation for Survitec Odoo exporter
# Reads stock_docs source files + live Odoo data and builds:
#   1. Sales report (6 Jul -> today)
#   2. Merged Nairobi front1+front2 physical count
#   3. Final expected stock per location
# All outputs written into stock_docs/
# =========================================================

import os
import re
import json
import xmlrpc.client
from datetime import datetime
from collections import defaultdict

import pandas as pd
from rapidfuzz import process, fuzz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

BASE = os.path.dirname(os.path.abspath(__file__))
DOCS = os.path.join(BASE, "stock_docs")

NBI = "NBI/Stock"   # Nairobi shop location (id 5)
RUI = "RUI/Stock"   # Ruiru warehouse location (id 14)

load_dotenv(os.path.join(BASE, ".env"))
ODOO_URL = os.getenv("ODOO_URL")
ODOO_DB = os.getenv("ODOO_DB")
ODOO_USERNAME = os.getenv("ODOO_USERNAME")
ODOO_PASSWORD = os.getenv("ODOO_API_KEY") or os.getenv("ODOO_PASSWORD")

# ---------------------------------------------------------
# ODOO ACCESS
# ---------------------------------------------------------
_common = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/common", allow_none=True)
_uid = _common.authenticate(ODOO_DB, ODOO_USERNAME, ODOO_PASSWORD, {})
if not _uid:
    raise SystemExit("AUTH FAILED")
_models = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/object", allow_none=True)
_FIELDS_CACHE = {}


def safe_call(model, method, args=None, kwargs=None, retries=3):
    if args is None:
        args = []
    if kwargs is None:
        kwargs = {}
    for attempt in range(retries):
        try:
            return _models.execute_kw(ODOO_DB, _uid, ODOO_PASSWORD, model, method, args, kwargs)
        except Exception as e:
            print(f"ODOO ERR {model}.{method} #{attempt+1}: {e}")
            import time
            time.sleep(2)
    return []


def get_valid_model_fields(model, preferred):
    if model not in _FIELDS_CACHE:
        try:
            _FIELDS_CACHE[model] = _models.execute_kw(
                ODOO_DB, _uid, ODOO_PASSWORD, model, "fields_get",
                [], {"attributes": ["name"]})
        except Exception:
            _FIELDS_CACHE[model] = None
    fm = _FIELDS_CACHE.get(model)
    if not fm:
        return preferred
    return [f for f in preferred if f in fm] or preferred


def search_read(model, fields, domain=None, limit=200000):
    return safe_call(model, "search_read", [domain or []], {"fields": fields, "limit": limit})


# ---------------------------------------------------------
# NAME NORMALISATION / MATCHING
# ---------------------------------------------------------
def norm(s):
    if s is None:
        return ""
    s = str(s)
    s = s.lower()
    s = s.replace("/", " ")
    s = re.sub(r"\(.*?\)", " ", s)         # strip bracketed qualifiers
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def disp_key(s):
    """Like norm but keeps the parenthetical variant qualifier (incl.
    spool weight) so every variant display stays distinct."""
    if s is None:
        return ""
    s = str(s).lower()
    s = re.sub(r"\s*\(\s*", "(", s)
    s = re.sub(r"\s*\)", ")", s)
    return re.sub(r"\s+", " ", s).strip()


def match_key(s):
    """Key for curated override lookups. Preserves parentheticals and
    plus signs so 'PLA+ Black' and 'PLA Black' stay distinct, unlike norm()."""
    if s is None:
        return ""
    s = str(s).lower()
    s = s.replace("+", " plus ")
    s = s.replace("/", " ")
    s = re.sub(r"[^a-z0-9(). ]+", " ", s)
    s = re.sub(r"\s*\(\s*", "(", s)
    s = re.sub(r"\s*\)", ")", s)
    return re.sub(r"\s+", " ", s).strip()


def to_num(v):
    try:
        if v is None or (isinstance(v, float) and v != v):
            return None
        return float(str(v).strip().replace(",", ""))
    except (ValueError, AttributeError):
        return None


class Matcher:
    """Fuzzy matches free-text product names against system variants."""
    def __init__(self, variants, overrides=None):
        self.variants = variants  # list of dicts with id, display, aliases
        self.by_norm = {}
        for v in variants:
            for alias in v["aliases"]:
                self.by_norm.setdefault(norm(alias), []).append(v["id"])
        self.overrides = overrides or {}
        self.lookup = [disp_key(v["display"]) for v in variants]
        self.ids = [v["id"] for v in variants]

    def match(self, name):
        n = norm(name)
        k = match_key(name)
        if k in self.overrides:
            return self.overrides[k], "override"
        if n in self.by_norm:
            return self.by_norm[n][0], "exact"
        # fuzzy search against display names (keeps variant qualifiers)
        best = process.extractOne(k, self.lookup, scorer=fuzz.QRatio,
                                  processor=lambda x: x, score_cutoff=72)
        if best:
            idx, score = best[2], best[1]
            return self.ids[idx], f"fuzzy:{score:g}"
        return None, "nomatch"


# ---------------------------------------------------------
# SYSTEM DATA FETCH
# ---------------------------------------------------------
def fetch_system():
    vf = get_valid_model_fields("product.product", ["id", "name", "display_name", "product_tmpl_id", "type", "lst_price", "standard_price"])
    tf = get_valid_model_fields("product.template", ["id", "name", "categ_id"])
    qf = get_valid_model_fields("stock.quant", ["product_id", "location_id", "quantity"])
    lf = get_valid_model_fields("stock.location", ["id", "name", "complete_name", "usage"])

    variants = search_read("product.product", vf)
    templates = {t["id"]: t for t in search_read("product.template", tf)}
    quants = search_read("stock.quant", qf)
    locations = {l["id"]: l for l in search_read("stock.location", lf)}

    # internal-only per-location stock
    internal = {lid for lid, l in locations.items() if l.get("usage") == "internal"}
    nbi_id = next((lid for lid, l in locations.items() if l.get("complete_name") == "NBI/Stock"), 5)
    rui_id = next((lid for lid, l in locations.items() if l.get("complete_name") == "RUI/Stock"), 14)

    loc_stock = defaultdict(lambda: defaultdict(float))
    for q in quants:
        pid = q.get("product_id") and q["product_id"][0]
        lid = q.get("location_id") and q["location_id"][0]
        if pid is None or lid is None:
            continue
        if lid not in internal:
            continue
        loc_stock[pid][lid] += float(q.get("quantity") or 0)

    return {"variants": variants, "templates": templates, "quants": quants,
            "locations": locations, "loc_stock": loc_stock, "nbi_id": nbi_id, "rui_id": rui_id}


def fetch_sales(from_date="2026-07-06"):
    plf = get_valid_model_fields("pos.order.line", ["id", "product_id", "qty", "price_subtotal", "order_id"])
    pof = get_valid_model_fields("pos.order", ["id", "date_order", "state"])
    slf = get_valid_model_fields("sale.order.line", ["id", "product_id", "product_uom_qty", "price_total", "order_id"])
    sof = get_valid_model_fields("sale.order", ["id", "date_order", "state"])
    vf = get_valid_model_fields("product.product", ["id", "display_name", "name", "product_tmpl_id"])
    tf = get_valid_model_fields("product.template", ["id", "name", "categ_id"])

    pos_lines = search_read("pos.order.line", plf)
    pos_orders = {o["id"]: o for o in search_read("pos.order", pof)}
    so_lines = search_read("sale.order.line", slf)
    so_orders = {o["id"]: o for o in search_read("sale.order", sof)}
    variants = {v["id"]: v for v in search_read("product.product", vf)}
    templates = {t["id"]: t for t in search_read("product.template", tf)}

    lines = []

    def fold(raw, orders, ok_states, qty_field, rev_field, channel):
        for line in raw:
            pid = line.get("product_id")
            if not pid:
                continue
            order = None
            oid = line.get("order_id")
            if oid:
                order = orders.get(oid[0])
            if not order or order.get("state") not in ok_states:
                continue
            dstr = str(order.get("date_order") or "")
            if not dstr or dstr[:10] < from_date:
                continue
            variant = variants.get(pid[0]) or {}
            tid = (variant.get("product_tmpl_id") or [None])[0]
            template = templates.get(tid) or {}
            lines.append({
                "date": dstr[:10], "product_id": pid[0], "template_id": tid,
                "product_name": variant.get("display_name") or variant.get("name") or "",
                "template_name": template.get("name") or "",
                "category": template.get("categ_id")[1] if template.get("categ_id") else "",
                "qty": float(line.get(qty_field) or 0),
                "revenue": float(line.get(rev_field) or 0),
                "channel": channel, "order_id": order["id"],
            })

    fold(pos_lines, pos_orders, ("paid", "done"), "qty", "price_subtotal", "POS")
    fold(so_lines, so_orders, ("sale",), "product_uom_qty", "price_total", "Sales Order")
    lines.sort(key=lambda x: x["date"])
    return lines


# ---------------------------------------------------------
# SOURCE FILE PARSERS
# ---------------------------------------------------------
def parse_consignment():
    """TSV content saved with .xlsx extension."""
    path = os.path.join(DOCS, "consignement.xlsx")
    rows = []
    with open(path, encoding="utf-8") as f:
        for i, line in enumerate(f):
            parts = line.rstrip("\n").split("\t")
            if i == 0:
                continue
            if len(parts) < 3:
                continue
            name = (parts[1] or "").strip()
            qty = to_num(parts[2]) if len(parts) > 2 else None
            if name and qty is not None:
                rows.append({"name": name, "qty": qty, "source": "Consignment"})
    return rows


def parse_from_warehouse():
    path = os.path.join(DOCS, "FROM-WAREHOUSE.xlsx")
    df = pd.read_excel(path, sheet_name="FROM-WAREHOUSE", header=0)
    rows = []
    for _, r in df.iterrows():
        name = str(r.get("PRODUCTS") or "").strip()
        qty = to_num(r.get("QTY"))
        if name and qty is not None:
            rows.append({"name": name, "qty": qty, "source": "FROM-WAREHOUSE"})
    return rows


def parse_pi():
    """Parse the latest proforma-invoice shipment file (pi*.xlsx, direct to
    Ruiru). Rows whose description cell is highlighted YELLOW mark goods
    that did NOT arrive - they are skipped entirely. Name = SKU + desc so
    bare colour words (e.g. TPU 'Black') stay unambiguous."""
    import glob
    files = sorted(glob.glob(os.path.join(DOCS, "pi*.xlsx")))
    if not files:
        print("WARNING: no pi*.xlsx found in stock_docs/")
        return []
    path = files[-1]
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Invoice"]
    rows = []
    for r in ws.iter_rows(min_row=11, max_row=60, max_col=8):
        sku = r[1].value
        desc = r[2].value
        qty = to_num(r[5].value)
        if not desc or qty is None:
            continue
        fill = r[2].fill
        yellow = (fill and fill.patternType == "solid"
                  and fill.fgColor and str(fill.fgColor.rgb) == "FFFFFF00")
        if yellow:
            continue
        name = f"{sku} {desc}".strip() if sku else str(desc).strip()
        name = name.replace("\uff08", "(").replace("\uff09", ")")
        rows.append({"name": name, "qty": qty, "source": "PI"})
    return rows


def parse_front1():
    """Parse NAIROBI_COUNT_FRONT1.xlsx. Rows 1-15 are filament groups
    (group/colour/qty), the rest are tools. Tool rows use several layouts:
      name|qty  |  group|variant|qty  |  (continuations follow a group header)"""
    path = os.path.join(DOCS, "NAIROBI_COUNT_FRONT1.xlsx")
    df = pd.read_excel(path, sheet_name="Sheet1", header=None)
    rows = []

    def cv(x):
        return "" if pd.isna(x) else str(x).strip()

    group = ""
    pending_base = ""
    for i, r in df.iterrows():
        a = cv(r[0])
        b = cv(r[1])
        c = r[2] if len(r) > 2 and pd.notna(r[2]) else None
        d = r[3] if len(r) > 3 and pd.notna(r[3]) else None
        cnum = to_num(c)
        dnum = to_num(d)
        bnum = to_num(b)

        if i < 16:  # filament block (rows 0..15)
            if a and b and cnum is not None:
                group = a
                rows.append({"name": f"{a} {b}", "qty": cnum, "source": "Front 1"})
            elif b and cnum is not None and group:
                rows.append({"name": f"{group} {b}", "qty": cnum, "source": "Front 1"})
            continue

        # tool block
        if a:
            if bnum is not None and cnum is None:
                # item name in col0, qty in col1
                rows.append({"name": a, "qty": bnum, "source": "Front 1"})
                pending_base = ""
            elif b and cnum is not None:
                # group header: group=col0, first variant=col1, qty=col2
                rows.append({"name": f"{a} {b}", "qty": cnum, "source": "Front 1"})
                pending_base = a
            elif cnum is not None and not b:
                rows.append({"name": a, "qty": cnum, "source": "Front 1"})
                pending_base = ""
            else:
                continue
        elif b and cnum is not None:
            if pending_base:
                rows.append({"name": f"{pending_base} {b}", "qty": cnum, "source": "Front 1"})
            else:
                rows.append({"name": b, "qty": cnum, "source": "Front 1"})
    return rows


def parse_front2():
    path = os.path.join(DOCS, "NAIROBI_COUNT_FRONT2.xlsx")
    df = pd.read_excel(path, sheet_name="Sheet1", header=0)
    rows = []
    for _, r in df.iterrows():
        name = (r.get("Product") or "").strip()
        if not name:
            continue
        qty = to_num(r.get("Counted at the shop"))
        q_sched = to_num(r.get("Quantity"))
        diff = to_num(r.get("Difference"))
        if qty is None:
            if q_sched is not None and diff is not None:
                qty = q_sched - diff
            elif q_sched is not None:
                qty = q_sched
            else:
                qty = 0.0
        rows.append({"name": name, "qty": qty, "source": "Front 2",
                     "location": r.get("Location")})
    return rows


def parse_warehouse():
    path = os.path.join(DOCS, "WAREHOUSE_STOCK.xlsx")
    df = pd.read_excel(path, sheet_name="Sheet1", header=None)
    rows = []

    def add(name, qty):
        name = " ".join(str(name).split())
        if name and qty is not None and not name.lower().startswith("item"):
            rows.append({"name": name, "qty": qty, "source": "Warehouse"})

    group_a = ""
    group_b = ""
    group_c = ""
    price_group = False
    for _, r in df.iterrows():
        vals = [r[j] if j < len(r) and pd.notna(r[j]) else None for j in range(22)]
        # Block A cols 0-3: group/color/qty
        a0 = str(vals[0]).strip() if vals[0] is not None else ""
        a1 = str(vals[1]).strip() if vals[1] is not None else ""
        a3 = to_num(vals[3])
        if a0 and a1 and a3 is not None:
            group_a = a0
            add(f"{a0} {a1}", a3)
        elif a0 and a3 is not None and not a1:
            add(a0, a3)
        elif a1 and a3 is not None:
            add(f"{group_a} {a1}", a3)

        # Block B cols 6-9: group/item/size/qty
        b6 = str(vals[6]).strip() if vals[6] is not None else ""
        b7 = str(vals[7]).strip() if vals[7] is not None else ""
        b8 = str(vals[8]).strip() if vals[8] is not None else ""
        b9 = to_num(vals[9])
        if b6:
            group_b = b6
        if b9 is not None and b7:
            name = " ".join(x for x in [group_b, b7, b8] if x)
            add(name, b9)

        # Block C cols 11-13: group/item/qty
        c11 = str(vals[11]).strip() if vals[11] is not None else ""
        c12 = str(vals[12]).strip() if vals[12] is not None else ""
        c13 = to_num(vals[13])
        if c11:
            group_c = c11
        if c13 is not None and c12:
            name = " ".join(x for x in [group_c, c12] if x)
            add(name, c13)

        # Block D cols 15-16: item/qty
        d15 = str(vals[15]).strip() if vals[15] is not None else ""
        d16 = to_num(vals[16])
        if d15 and d16 is not None:
            add(d15, d16)

        # Block E cols 18-21: Price labels color/qty
        e18 = str(vals[18]).strip() if vals[18] is not None else ""
        e19 = str(vals[19]).strip() if vals[19] is not None else ""
        e21 = to_num(vals[21])
        if e18:
            price_group = e18.lower().startswith("price label")
        if e21 is not None and e19:
            if price_group:
                add(f"Price Labels {e19}", e21)
            else:
                add(e19, e21)
    return rows


# ---------------------------------------------------------
# MATCH + RECONCILE
# ---------------------------------------------------------
NON_SALEABLE = {
    "[Delivery_007] Standard delivery", "[COMM] Communication",
    "[EXP_GEN] Expenses", "[FOOD] Meals", "[GIFT] Gifts",
    "[MIL] Mileage", "[TIPS] Tips", "[TRANS & ACC] Travel & Accommodation",
    "Deposit", "Down Payment (POS)", "Pick up in store",
    "Settle Due", "Settle Invoice", "Shipping fee", "Survi Delivery",
}


def build_variant_objects(variants, templates):
    out = []
    for v in variants:
        vtype = v.get("type") or ""
        disp = (v.get("display_name") or v.get("name") or "").strip()
        if disp in NON_SALEABLE:
            continue
        tid = v.get("product_tmpl_id") and v["product_tmpl_id"][0]
        tname = (templates.get(tid) or {}).get("name") or ""
        aliases = {disp, tname, v.get("name") or ""}
        aliases.add(re.sub(r",\s*spool weight:?\s*[^)]*", "", disp, flags=re.I))
        m = re.match(r"^(.*?)\s*\((.*)\)\s*$", disp)
        if m:
            base = m.group(1).strip()
            color = m.group(2).split(",")[0].strip()
            aliases.add(f"{base} {color}")
            aliases.add(f"{color} {base}")
            aliases.add(color)
            # color without the material token repeated, e.g. "Bright Red PLA +" -> "Bright Red"
            words = color.split()
            for cut in range(len(words), 0, -1):
                cand = " ".join(words[:cut])
                if norm(cand) not in norm(color):
                    break
                aliases.add(cand)
        out.append({
            "id": v["id"], "display": disp, "template": tname,
            "type": vtype, "aliases": [x for x in aliases if x],
            "buy": v.get("standard_price"), "sell": v.get("lst_price"),
        })
    return out


def aggregate(rows):
    agg = defaultdict(float)
    for r in rows:
        agg[r["name"]] += r["qty"]
    return agg


# ---------------------------------------------------------
# CURATED OVERRIDES: manual shorthand -> system variant
# ---------------------------------------------------------
# Each entry: (manual name as parsed, system display name to map to).
# Looked up at runtime against the live variant list so ids never go stale.
_OVERRIDE_TABLE = [
    # ---- Filaments: Front1 groups ----
    ("PLA+ Bright red", "PLA + (Bright Red PLA +)"),
    ("PLA+ Bright skin", "PLA + (Bright Skin PLA +)"),
    ("PLA+ Golden", "PLA + (Golden PLA +)"),
    ("PLA+ Grey", "PLA + (Grey PLA+)"),
    ("PLA+ Purple", "PLA + (Purple PLA+)"),
    ("PLA+ Orange", "PLA + (Orange PLA+)"),
    ("PLA+ Red", "PLA + (Red PLA+)"),
    ("PLA+ Silver", "PLA + (Silver PLA+)"),
    ("PLA+ Skin", "PLA + (Skin PLA +)"),
    ("PLA+ White", "PLA + (White PLA +)"),
    ("PLA+ Blue", "PLA + (Blue PLA +)"),
    ("PLA+ Black", "PLA + (Black PLA +)"),
    ("PLA+ Green", "PLA + (Green PLA +)"),
    ("PLA+ Yellow", "PLA + (Yellow PLA +)"),
    ("PLA+ Cyan", "PLA + (cyan)"),
    ("PLA+ Transparent white", "PLA + (Transparent White PLA +)"),
    ("PLA Yellow", "PLA Basic (Yellow PLA)"),
    ("PLA Golden", "PLA Basic (Golden PLA Basic)"),
    ("PLA Silver", "PLA Basic (Silver PLA Basic)"),
    ("PLA Skin", "PLA Basic (Skin PLA Basic)"),
    ("PLA Purple", "PLA Basic (Purple PLA Basic)"),
    ("PLA Marble", "PLA Basic (Marble PLA Basic)"),
    ("PLA White", "PLA Basic (White PLA Basic)"),
    ("PLA Dark grey", "PLA Basic (Dark Grey PLA Basic)"),
    ("PLA Blue", "PLA Basic (Blue PLA)"),
    ("PLA Black", "PLA Basic (Black PLA Basic)"),
    ("PLA Red", "PLA Basic (Red PLA Basic)"),
    ("PLA Green", "PLA Basic (Green PLA Basic)"),
    ("PLA Sky blue", "PLA Basic (Sky Blue PLA Basic)"),
    ("PLA Brown", "PLA Basic (Brown PLA Basic)"),
    ("PLA Orange", "PLA Basic (Orange PLA Basic)"),
    ("PLA Mint green", "PLA Basic (Mint Green)"),
    ("PLA Grey", "PLA Basic (Grey PLA Basic)"),
    ("PLA Transparent", "PLA Basic (Transparent PLA Basic)"),
    ("PLA Woodiness", "PLA Basic (Woodiness PLA Basic)"),
    ("PLA Blue to white", "PLA Basic (Blue to White PLA Basic (Blue when cold White when hot))"),
    ("PLA Luminous pink", "Luminous PLA Basic (Luminous Pink PLA Basic)"),
    ("PLA Luminous Blue", "Luminous PLA Basic (Luminous Blue PLA Basic)"),
    ("PLA Luminous yellow", "Luminous PLA Basic (Luminous Yellow PLA Basic)"),
    ("PLA Luminous green", "Luminous PLA Basic (Luminous Green PLA Basic)"),
    ("PLA Luminous purple", "Luminous PLA Basic (Luminous Purple BLA Basic)"),
    ("PLA Matte red", "Matte PLA Filament (Matte Red PLA)"),
    ("PLA Matte Blue", "Matte PLA Filament (Matte Blue PLA)"),
    ("PLA Matte Yellow", "Matte PLA Filament (Matte Yellow PLA)"),
    ("PLA Matte Purple", "Matte PLA Filament (Matte Purple PLA)"),
    ("PLA Matte grass green", "Matte PLA Filament (Matte Grass Green PLA)"),
    ("PLA Matte Green", "Matte PLA Filament (Matte Green PLA)"),
    ("PLA Matte Black", "Matte PLA Filament (Matte Black PLA)"),
    ("PLA Matte White", "Matte PLA Filament (Matte White PLA)"),
    ("PLA Matte Grey", "Matte PLA Filament (Matte Grey PLA)"),
    ("PLA Matte Orange", "Matte PLA Filament (Matte Orange PLA)"),
    ("PLA Matte Brown", "Matte PLA Filament (Matte Brown PLA)"),
    ("PLA Matte Skin", "Matte PLA Filament (Matte Skin PLA)"),
    ("PLA Matte Shade brown", "Matte PLA Filament (Matte Shade Brown)"),
    ("PLA Matte Rainbow", "Matte PLA Filament (Matte Rainbow PLA)"),
    ("PLA Matte Lillac purple", "Matte PLA Filament (Matte Lillac Purple PLA)"),
    ("Silk Silver", "Silk PLA (Silk Silver)"),
    ("Silk Gold", "Silk PLA (Silk Golden)"),
    ("Silk Copper", "Silk PLA (Silk Copper PLA)"),
    ("Silk Orange", "Silk PLA (Silk Orange)"),
    ("Silk Purple", "Silk PLA (Silk Purple PLA)"),
    ("Silk Green", "Silk PLA (Silk Green PLA)"),
    ("Silk Red", "Silk PLA (Silk Red PLA)"),
    ("Silk Yellow", "Silk PLA (Silk Yellow)"),
    ("Silk Blue", "Silk PLA (Silk Blue PLA)"),
    ("Silk Black", "Silk PLA (Silk Black PLA)"),
    ("Silk Glow rainbow", "Silk Glow Rainbow"),
    ("Silk Gradient rainbow", "Silk Gradient Rainbow"),
    ("Silk Rainbow macaron", "Silk Rainbow Macaron"),
    ("Silk Rainbow universe", "Silk Rainbow Universe Series"),
    ("Matte PLA Green", "Matte PLA Filament (Matte Green PLA)"),
    ("Matte PLA Grass green", "Matte PLA Filament (Matte Grass Green PLA)"),
    ("Matte PLA Grey", "Matte PLA Filament (Matte Grey PLA)"),
    ("Matte PLA Blue", "Matte PLA Filament (Matte Blue PLA)"),
    ("Matte PLA Red", "Matte PLA Filament (Matte Red PLA)"),
    ("Matte PLA Orange", "Matte PLA Filament (Matte Orange PLA)"),
    ("Matte PLA Yellow", "Matte PLA Filament (Matte Yellow PLA)"),
    ("Matte PLA Black", "Matte PLA Filament (Matte Black PLA)"),
    ("Matte PLA White", "Matte PLA Filament (Matte White PLA)"),
    ("Matte PLA Purple", "Matte PLA Filament (Matte Purple PLA)"),
    ("Matte PLA Brown", "Matte PLA Filament (Matte Brown PLA)"),
    ("Matte PLA Skin", "Matte PLA Filament (Matte Skin PLA)"),
    ("Matte PLA Shade brown", "Matte PLA Filament (Matte Shade Brown)"),
    ("ABS White", "ABS Filament (White ABS)"),
    ("ABS Pink", "ABS Filament (Pink ABS)"),
    ("ABS Black", "ABS Filament (Black ABS)"),
    ("ABS Blue", "ABS Filament (Blue ABS)"),
    ("ABS Yellow", "ABS Filament (ABS Yellow)"),
    ("ABS Grey", "ABS Filament (Grey ABS)"),
    ("ABS Orange", "ABS Filament (Orange ABS)"),
    ("ABS Red", "ABS Filament (Red ABS)"),
    ("ABS Green", "ABS Filament (Green ABS)"),
    ("ABS Silver", "ABS Filament (ABS Silver)"),
    ("TPU Transparent", "TPU Filaments (Transparent TPU)"),
    ("TPU White", "TPU Filaments (White TPU)"),
    ("TPU Black", "TPU Filaments (Black TPU)"),
    ("TPU Blue", "TPU Filaments (Blue TPU)"),
    ("TPU Green", "TPU Filaments (Green TPU)"),
    ("TPU Yellow", "TPU Filaments (Yellow TPU)"),
    ("TPU Red", "TPU Filaments (Red TPU)"),
    ("TPU Orange", "TPU Filaments (Orange TPU)"),
    ("TPU Grey", "TPU Filaments (Grey TPU)"),
    ("Marble PLA", "Marble PLA Basic"),
    ("CF Black Nylon", "(PACF) Carbon Fiber Nylon 3D Filament"),
    ("CF Black Carbon", "PLA Carbon Fiber Basic"),
    ("Black Carbon", "PLA Carbon Fiber Basic"),

    # ---- Warehouse filament blocks (group already prepended) ----
    ("PETG Florescent Yellow", "PETG (Fluorescent Yellow PETG, Spool Weight: 1Kg)"),
    ("PETG Florescent Red", "PETG (Fluorescent Red PETG, Spool Weight: 1Kg)"),
    ("PETG Florescent Blue", "PETG (Fluorescent Blue PETG, Spool Weight: 1Kg)"),
    ("PETG Pink", "PETG (Pink PETG, Spool Weight: 1Kg)"),
    ("PETG Silver", "PETG (Silver PETG, Spool Weight: 1Kg)"),
    ("PETG Metallic Silver", "PETG (Metallic Silver PETG, Spool Weight: 1Kg)"),
    ("PETG Metal Copper", "PETG (Metal Copper PETG, Spool Weight: 1Kg)"),
    ("PETG Red", "PETG (Red PETG, Spool Weight: 1Kg)"),
    ("PETG Blue", "PETG (Blue PETG, Spool Weight: 1Kg)"),
    ("PETG White", "PETG (White PETG, Spool Weight: 1Kg)"),
    ("PETG Black", "PETG (Black PETG, Spool Weight: 1Kg)"),
    ("PETG Green", "PETG (Green PETG, Spool Weight: 1Kg)"),
    ("PETG Grass Green", "PETG (Grass Green PETG, Spool Weight: 1Kg)"),
    ("PETG Grey", "PETG (Grey PETG, Spool Weight: 1Kg)"),
    ("PETG Transparent", "PETG (Transparent PETG, Spool Weight: 1Kg)"),
    ("PETG Yellow", "PETG (Yellow PETG, Spool Weight: 1Kg)"),
    ("PETG Sky Blue", "PETG (Sky Blue PETG, Spool Weight: 1Kg)"),
    ("PETG Skin", "PETG (Skin PETG, Spool Weight: 1Kg)"),
    ("PETG Orange", "PETG (Orange PETG, Spool Weight: 1Kg)"),
    ("PETG Dark Blue", "PETG (Dark Blue PETG, Spool Weight: 1Kg)"),
    ("PETG Mint Green", "PETG (Mint Green PETG, Spool Weight: 1Kg)"),
    ("PETG Light Pink", "PETG (Light Pink PETG, Spool Weight: 1Kg)"),
    ("PETG Light Apricot", "PETG (Light Apricot PETG, Spool Weight: 1Kg)"),
    ("PETG White 3kg", "PETG (White) 3kg"),
    ("PETG Black 3kg", "PETG (Black) 3kg"),
    ("Silk tri Red Yellow Blue", "Silk Multi-Color PLA (Silk Red Yellow Blue)"),
    ("Silk tri Red Green Blue", "Silk Multi-Color PLA (Silk Red Green Blue)"),
    ("Silk tri Red Gold Purple", "Silk Multi-Color PLA (Silk Red Gold Purple)"),
    ("Silk tri Gold green black", "Silk Multi-Color PLA (Silk Gold Green Black)"),
    ("Silk tri Black blue purple", "Silk Multi-Color PLA (Silk Black Blue Purple)"),
    ("Silk tri Yellow blue green", "Silk Multi-Color PLA (Silk Yellow Blue Green)"),
    ("Silk tri Green purplecopper", "Silk Multi-Color PLA (Silk Green Purple Copper)"),
    ("Silk tri Blue green purple", "Silk Multi-Color PLA (Silk Blue Green Purple)"),
    ("Silk tri Red gold blue", "Silk Multi-Color PLA (Silk Red Gold Blue)"),
    ("Silk tri Gold green blue", "Silk Multi-Color PLA (Silk Gold Green Blue)"),
    ("Silk tri Gold purple red black", "Silk Multi-Color PLA (Silk Gold Purple Red Black)"),
    ("Silk tri Gold silver copper", "Silk Multi-Color PLA (Silk Gold Silver Copper)"),
    ("Silk tri Gradient candy", "Silk Multi-Color PLA (Silk Gradient Candy)"),
    ("Silk tri Gradient macaron", "Silk Multi-Color PLA (Silk Gradient Macaron)"),
    ("Silk tri Ice cream", "Silk Multi-Color PLA (Silk Ice Cream Tones)"),
    ("Silk tri Red gold green", "Silk Multi-Color PLA (Silk Red Gold Green)"),
    ("Bi-colour Gold Purple", "Silk Bi-Color PLA (Silk Gold Purple PLA)"),
    ("Bi-colour Red Gold", "Silk Bi-Color PLA (Silk Red Gold)"),
    ("Bi-colour Pink Gold", "Silk Bi-Color PLA (Silk Pink Gold)"),
    ("Bi-colour Blue green", "Silk Bi-Color PLA (Silk Blue Green PLA)"),
    ("Bi-colour Green black", "Silk Bi-Color PLA (Silk Green Black PLA)"),
    ("Bi-colour Blue Purple", "Silk Bi-Color PLA (Silk Blue Purple PLA)"),
    ("Bi-colour Red blue", "Silk Bi-Color PLA (Silk Red Blue PLA)"),
    ("Bi-colour Red Green", "Silk Bi-Color PLA (Silk Red Green PLA)"),
    ("Bi-colour Gold Silver", "Silk Bi-Color PLA (Silk Gold Silver PLA)"),
    ("Bi-colour Yellow Green", "Silk Bi-Color PLA (Silk Yellow Green PLA)"),
    ("Bi-colour Red Black", "Silk Bi-Color PLA (Silk Red Black PLA)"),
    ("Silk Orange", "Silk PLA (Silk Orange)"),
    ("Silk Purple", "Silk PLA (Silk Purple PLA)"),
    ("Silk Yellow", "Silk PLA (Silk Yellow)"),
    ("Silk Green", "Silk PLA (Silk Green PLA)"),
    ("Silk Red", "Silk PLA (Silk Red PLA)"),
    ("Silk Golden", "Silk PLA (Silk Golden)"),
    ("Silk Copper", "Silk PLA (Silk Copper PLA)"),
    ("Silk Black", "Silk PLA (Silk Black PLA)"),
    ("Silk Blue", "Silk PLA (Silk Blue PLA)"),
    ("Silk Silver", "Silk PLA (Silk Silver)"),
    ("Multicolour Purple red blue green", "Silk Multi-Color PLA (Silk Purple Red Blue Green)"),
    ("Multicolour Gold green rose red", "Silk Multi-Color PLA (Silk Gold Green Rose Red)"),
    ("Multicolour Gradient rainbow", "Silk Gradient Rainbow"),
    ("Multicolour Gold purple red blue", "Silk Multi-Color PLA (Silk Gold Purple Red Blue)"),
    ("Multicolour Black blue purple", "Silk Multi-Color PLA (Silk Black Blue Purple)"),
    ("Multicolour Blue green orange", "Silk Multi-Color PLA (Silk Blue Green Orange)"),
    ("Antinsky PLA Green", "PLA Basic (Green PLA Basic)"),
    ("Antisky PLA Black", "PLA Basic (Black PLA Basic)"),
    ("Antinsky PLA White", "PLA Basic (White PLA Basic)"),
    ("Antinsky PETG Transparent", "PETG (Transparent PETG, Spool Weight: 1Kg)"),
    ("Antinsky PETG White", "PETG (White PETG, Spool Weight: 1Kg)"),
    ("Antinsky PETG Grey", "PETG (Grey PETG, Spool Weight: 1Kg)"),
    ("Polymaker Matte fossil grey", "polymakers (fossil grey)"),
    ("Polymaker Matte cotton white", "polymakers (cotton white)"),
    ("Refill Matte Cotton White", "polymakers (cotton white)"),
    ("Ender 3V3 KE", "Creality Ender-3 V3 3D Printer (Model: KE)"),

    # ---- Warehouse welding / gas / tools ----
    ("Welding rods AWS E308L 2.5", "Arch Welding Rods (AWS E308L-16, 2.5)"),
    ("Welding rods AWS E308L 3.2", "Arch Welding Rods (AWS E308L-16, 3.2)"),
    ("Welding rods AWS E7018 2.5", "Arch Welding Rods (AWS E7018, 2.5)"),
    ("Welding rods AWS E7018 3.2", "Arch Welding Rods (AWS E7018, 3.2)"),
    ("Welding rods AWS E6013 2.5", "Arch Welding Rods (AWS E6013, 2.5)"),
    ("Welding rods AWS E6013 3.2", "Arch Welding Rods (AWS E6013, 3.2)"),
    ("Mig wires ER70-S 5kg 0.8", "Mig Wire for Gas (Wire Diameter: 0.8mm, Weight: 5kg)"),
    ("Mig wires ER70-S 5kg 1kg", "Mig Wire for Gas (Wire Diameter: 1.0mm, Weight: 5kg)"),
    ("Mig wires ER70-S 1KG 1mm", "Mig Wire for Gas (Wire Diameter: 1.0mm, Weight: 1kg)"),
    ("Mig wires ER70-S 1KG 0.8mm", "Mig Wire for Gas (Wire Diameter: 0.8mm, Weight: 1kg)"),
    ("Mig wires E71T-GS 5Kg 1mm", "Flux Cored Mig Wire (Gasless) (Weight: 5kg, Wire Diameter: 1.0mm)"),
    ("Mig wires E71T-GS 5Kg 0.8mm", "Flux Cored Mig Wire (Gasless) (Weight: 5kg, Wire Diameter: 0.8mm)"),
    ("Mig wires E71T-GS 1KG 0.8mm", "Flux Cored Mig Wire (Gasless) (Weight: 1kg, Wire Diameter: 0.8mm)"),
    ("Mig wires E71T-GS 1KG 1.0mm", "Flux Cored Mig Wire (Gasless) (Weight: 1kg, Wire Diameter: 1.0mm)"),
    ("Mig wires E71T-GS 1KG 1.0 mm", "Flux Cored Mig Wire (Gasless) (Weight: 1kg, Wire Diameter: 1.0mm)"),
    ("Mig wires ER70 -S 1 KG 1mm", "Mig Wire for Gas (Wire Diameter: 1.0mm, Weight: 1kg)"),
    ("Mig wires ER70 -S 1 KG 0.8 mm", "Mig Wire for Gas (Wire Diameter: 0.8mm, Weight: 1kg)"),
    ("Mig wires ER70-S 1KG 0.8 mm", "Mig Wire for Gas (Wire Diameter: 0.8mm, Weight: 1kg)"),
    ("Precise drill stand (green)", "Portable Electric Drill Holder Stand"),
    ("Helmets Grafity", "Autodarkening Welding Mask Helmet (Gravity)"),
    ("Gas Argon 40l", "Argon Gas Cylinder with Gas(40 Liters)"),
    ("Argon 14l", "Argon cylinder with gas (14L)"),
    ("Regulators co2", "CO2 Gas Regulator"),
    ("Regulators argon male", "Argon Regulator male"),
    ("Regulators argon female", "Argon regulator female"),
    ("Plastic dowel", "Plastic Dowel"),
    ("Pipe clamp 3/4", "3/4'' Pipe Clamp without stand"),
    ("Pipe clamp with stand", "3/4\" Pipe Clamp With Stand"),
    ("1/2\" pipe clamp", "1/2\" pipe clamps"),
    ("Corner clamps 68mm", "Angle Corner Clamps"),
    ("Angle grinder stand (black)", "4.5 Angle Grinder stand (black)"),
    ("Electric drill stand", "Portable Electric Drill Holder Stand"),
    ("Drill press vice", "Drill press vice"),
    ("7'' Speed square", "7\" Speed Square"),
    ("12'' Speed square", "12\" speed square"),
    ("30 cm steel ruler", "30cm steel ruler"),
    ("Pipe bender", "2\" Hydraulic Pipe Bender"),
    ("Price Labels Luminous green", "Price Gun Label Sticker Bright Green"),
    ("Price Labels Dark Orange", "Price Gun Label Sticker Dark Orange"),
    ("Price Labels Light Orange", "Price Gun Label Sticker Bright Orange"),
    ("Price Labels Bright Orange", "Price Gun Label Sticker Bright Orange"),
    ("Price Labels Blue", "Price Gun Label Sticker Blue"),
    ("Price Labels White", "Price Gun Label Sticker White"),
    ("Price Labels Dark green", "Price Gun Label Sticker Dark Green"),
    ("Price Labels Purple", "Price Gun Label Sticker Purple"),
    ("Price Labels Pink", "Price Gun Label Sticker Pink"),
    ("Price Labels Bright Green", "Price Gun Label Sticker Bright Green"),
    ("Price Labels Light green", "Price Gun Label Sticker Bright Green"),

    # ---- Warehouse lathe chucks (K-series -> size) ----
    ("Chucks K11250 3 J Self c", "lathe Chucks (self centering 3jaw, 10\")"),
    ("Chucks K11200 3 J self c", "lathe Chucks (self centering 3jaw, 8\")"),
    ("Chucks K21250 3 J self c", "lathe Chucks (self centering 3jaw, 10\")"),
    ("Chucks K11 200 3 J self c", "lathe Chucks (self centering 3jaw, 8\")"),
    ("Chucks K21 100 3 J Self c", "lathe Chucks (self centering 3jaw, 4\")"),
    ("Chucks K11 160 /D4 3 J Self c", "lathe Chucks (self centering 3jaw, 6\")"),
    ("Chucks K12 160 2 J Self c", "lathe Chucks (self centering 2jaw, 6\")"),
    ("Chucks K10 100 2 J Self c", "lathe Chucks (self centering 2jaw, 4\")"),
    ("Chucks K12 100 4 J self c", "lathe Chucks (self centering 4jaw, 4\")"),
    ("Chucks K12 200 4 J self c", "lathe Chucks (self centering 4jaw, 8\")"),
    ("Chucks K11 200/A26 3 J Self c", "lathe Chucks (self centering 3jaw, 8\")"),
    ("Chucks K12 160 4 J self c", "lathe Chucks (self centering 4jaw, 6\")"),
    ("Chucks K10 200 2 J Self c", "lathe Chucks (self centering 2jaw, 8\")"),
    ("Chucks K10 250 2 J Self c", "lathe Chucks (self centering 2jaw, 10\")"),
    ("Chucks K11 320 3 J Self c", "lathe Chucks (self centering 3jaw, 12\")"),

    # ---- LATHE HOLDERS (ISO codes -> system holder type/size) ----
    ("LATHE HOLDERS MGHR 2020-4", "Lathe Tool Holders (Roughing Tools, 20mm)"),
    ("LATHE HOLDERS SCLCR 1212 HO6", "Lathe Tool Holders (Turning Tool, 12mm)"),
    ("LATHE HOLDERS SDJCR 2020K11", "Lathe Tool Holders (Turning Tool, 20mm)"),
    ("LATHE HOLDERS SDJCR 2525M11", "Lathe Tool Holders (Turning Tool, 25mm)"),
    ("LATHE HOLDERS SNR-0025S16", "Lathe Tool Holders (Boring Tool, 25mm)"),
    ("LATHE HOLDERS S25SSCLCL09", "Lathe Tool Holders (Boring Tool, 25mm)"),
    ("LATHE HOLDERS S25SCLCR09", "Lathe Tool Holders (Boring Tool, 25mm)"),
    ("LATHE HOLDERS SDJCR 1212H07", "Lathe Tool Holders (Turning Tool, 12mm)"),
    ("LATHE HOLDERS SER 1212 H16", "Lathe Tool Holders (External Threading Tools, 12mm)"),
    ("LATHE HOLDERS MGEHR 2525 2", "Lathe Tool Holders (Roughing Tools, 25mm)"),
    ("LATHE HOLDERS SNR0020R16", "Lathe Tool Holders (Boring Tool, 20mm)"),
    ("LATHE HOLDERS MGEHR 1212 2", "Lathe Tool Holders (Roughing Tools, 12mm)"),
    ("LATHE HOLDERS SCLCR 2020 K09", "Lathe Tool Holders (Turning Tool, 20mm)"),
    ("LATHE HOLDERS SCLCR 2525 M09", "Lathe Tool Holders (Turning Tool, 25mm)"),
    ("LATHE HOLDERS SER 20220 K16", "Lathe Tool Holders (External Threading Tools, 20mm)"),
    ("LATHE HOLDERS S20RSCLCR9", "Lathe Tool Holders (Boring Tool, 20mm)"),
    ("LATHE HOLDERS SNR0012M11", "Lathe Tool Holders (Boring Tool, 12mm)"),
    ("LATHE HOLDERS MGHR 2020-2", "Lathe Tool Holders (Roughing Tools, 20mm)"),
    ("LATHE HOLDERS S20R-SCLCL09", "Lathe Tool Holders (Boring Tool, 20mm)"),
    ("LATHE HOLDERS SNR 0012M11", "Lathe Tool Holders (Boring Tool, 12mm)"),
    ("LATHE HOLDERS SER 2525 M16", "Lathe Tool Holders (External Threading Tools, 25mm)"),
    ("LATHE HOLDERS SDJCR 1616 H11", "Lathe Tool Holders (Turning Tool, 16mm)"),
    ("LATHE HOLDERS SDNCN 1212 H07", "Lathe Tool Holders (Turning Tool, 12mm)"),
    ("LATHE HOLDERS SCLCL 1616 H09", "Lathe Tool Holders (Turning Tool, 16mm)"),
    ("LATHE HOLDERS SCLCL 2020 K09", "Lathe Tool Holders (Turning Tool, 20mm)"),
    ("LATHE HOLDERS SNL00 20R16", "Lathe Tool Holders (Boring Tool, 20mm)"),
    ("LATHE HOLDERS S20R SCLCR09", "Lathe Tool Holders (Boring Tool, 20mm)"),
    ("LATHE HOLDERS S16Q SDX CR11", "Lathe Tool Holders (Boring Tool, 16mm)"),
    ("LATHE HOLDERS SDNCN 1616 H11", "Lathe Tool Holders (Turning Tool, 16mm)"),
    ("LATHE HOLDERS SDNCN 2020 K11", "Lathe Tool Holders (Turning Tool, 20mm)"),
    ("LATHE HOLDERS SEL 1212 H16", "Lathe Tool Holders (External Threading Tools, 12mm)"),
    ("LATHE HOLDERS SNL0012 M11", "Lathe Tool Holders (Boring Tool, 12mm)"),
    ("LATHE HOLDERS SDJCL 1212 H07", "Lathe Tool Holders (Turning Tool, 12mm)"),
    ("LATHE HOLDERS SDJCR 1212 H07", "Lathe Tool Holders (Turning Tool, 12mm)"),
    ("LATHE HOLDERS s16q -sclcr09", "Lathe Tool Holders (Boring Tool, 16mm)"),

    # ---- Other wrong fuzzy matches corrected ----
    ("Welding rods E7018 2.5", "Arch Welding Rods (AWS E7018, 2.5)"),
    ("Gas Argon 14l", "Argon cylinder with gas (14L)"),
    ("1.0mm/1kg Copper Cored Mig Wire", "Mig Wires (gas, Wire Diameter: 1.0mm, Weight: 1kg)"),
    ("0.8mm/1kg Copper Cored Mig Wire", "0.8mm/1kg Copper Coated Mig Wire"),
    ("Aluminium 3D Ruler Blue", "Aluminium 3D Red"),
    ("Aluminium 3D Ruler Black", "Aluminium 3D Red"),
    ("Portable Electric Drill silver", None),
    ("Pocket hole screws ST4- 38", None),
    ("Bench drill square (green)", None),
    ("3\" G-CLAMP", None),
    ("4\" G-CLAMP", None),
    ("6\" G-CLAMP", None),
    ("4PC F-CLAMP SET", None),

    # ---- Front1 stragglers ----
    ("Glass cutter with opener set", "2 in 1 Glass and Tile Cutter"),
    ("10 PC Solid carbide endmill (blue coating)", None),
    ("Saw chain filing guide", None),
    ("3mm carbide engraving burr", None),
    ("smoothing planer", "235x50 mm planner"),
    ("Planner ( h treimmer)", "235x50 mm planner"),
    ("steel roteey burrs(assorted chisel)", "Carbide Rotaryy Burs 6mm Shank"),
    ("inserts CCMTO60204 VP15", "Carbide Inserts (CCMT060204 VP15)"),
    ("telescopic quick support rod", None),
    ("1/8 rotary burrs", "10 pc Tungsten Rotary burrs"),
    ("pipe", "Gas pipe"),
    ("auger bits 8pc incomplete", "Auger Bits  Set"),
    ("HS PLA Cayan", "PLA + (cyan)"),
    ("PLA Basic Silk Yellow Green 3D Printing Filament", "Silk Bi-Color PLA (Silk Yellow Green PLA)"),
    ("HSS Square Twist Drill Bits 6.4mm", "4PCS HSS Twist Drill Bits"),
    ("HSS Square Twist Drill Bits 9.5mm", "4PCS HSS Twist Drill Bits"),
    ("HSS Square Twist Drill Bits 11mm", "4PCS HSS Twist Drill Bits"),
    ("HSS Square Twist Drill Bits 12.7mm", "4PCS HSS Twist Drill Bits"),
    ("HSS Square Twist Drill Bits 14mm", "4PCS HSS Twist Drill Bits"),
    ("ER-32 10mm Collet", "ER-32 Collet (10mm)"),
    ("ER-32 13mm Collet", "ER-32 Collet (13mm)"),
    ("ER-32 4mm Collet", "ER-32 Collet (4mm)"),
    ("ER-32 5mm Collet", "ER-32 Collet (5mm)"),
    ("ER-32 6mm Collet", "ER-32 Collet (6mm)"),
    ("ER-32 16mm Collet", "ER-32 Collet (16mm)"),
    ("ER-32 17mm Collet", "ER-32 Collet (17mm)"),
    ("ER-32 18mm Collet", "ER-32 Collet (18mm)"),
    ("ER-32 19mm Collet", "ER-32 Collet (19mm)"),
    ("ER-32 20mm Collet", "ER-32 Collet (20mm)"),
    ("Collet (7mm)", "ER-32 Collet (7mm)"),
    ("Collet (11mm)", "ER-32 Collet (11mm)"),
    ("Collet (13mm)", "ER-32 Collet (13mm)"),
    ("Collet (14mm)", "ER-32 Collet (14mm)"),
    ("Collet (15mm)", "ER-32 Collet (15mm)"),
    ("ASA Gray 3D Printing Filament", "ASA 3D Printing Filament (ASA Grey)"),
    ("ASA White 3D Printing Filament", "ASA 3D Printing Filament (ASA White)"),

    # ---- Front2 stragglers ----
    ("C Clamp 11", None),
    ("C Clamp 14", None),
    ("C Clamp 18", None),

    # ---- Front1 tools ----
    ("Digital depth gauge", "Digital Depth Gauge"),
    ("Mini diamond blades", "Mini Diamond Saw Bades"),
    ("Tap and die set (32 pcs)", "32pcs tap and die set"),
    ("Hss saw blades", "6 pc Hss saw blades"),
    ("Autodarkening welding googles", "Auto Darkening Welding Goggles"),
    ("Drill press vice", "Drill press vice"),
    ("2 in 1 Scriber", "2 in 1 Scriber"),
    ("Dial gauge", "Dial Gauge"),
    ("Nylon clamps size 2", "2\" Nylon clamps"),
    ("Nylon clamps size 3", "3\" Nylon Clamps"),
    ("Nylon clamps size 3.5", "Nylon clamps 3.5"),
    ("Nylon clamps size 6", "6\" Nylon Clamps"),
    ("Digital plastic caliper", "Plastic Digital Vernier Caliper"),
    ("Contour gauge", "25 mm contour gauge"),
    ("Alloy scribe pen", "Scribing pens"),
    ("Woodworking angle ruler", "Woodworking angle ruler"),
    ("90 corner clamp", "90Â° Single Jaw Corner Clamp"),
    ("Mini Planer (heavy duty)", "Mini Planer (Middle Heavy)"),
    ("Mini Planer (light duty)", "Mini Planer (Middle light weight)"),
    ("Wood hole saw set (5 pc)", "5pc Hole saw set"),
    ("Malfa bits (red)", "Malfa bits(Hole Saw Set)"),
    ("15 pcs router bit set 6mm", "15Pc Router Bit Set"),
    ("15 pcs router bit set 8mm", "15Pc Router Bit Set"),
    ("Price labels Dark green", "Price Gun Label Sticker Dark Green"),
    ("Price labels Light green", "Price Gun Label Sticker Bright Green"),
    ("Price labels Bright Orange", "Price Gun Label Sticker Bright Orange"),
    ("Price labels Dark Orange", "Price Gun Label Sticker Dark Orange"),
    ("Price labels Purple", "Price Gun Label Sticker Purple"),
    ("Price labels White", "Price Gun Label Sticker White"),
    ("Price labels Blue", "Price Gun Label Sticker Blue"),
    ("Price labels Pink", "Price Gun Label Sticker Pink"),
    ("92 PC Dremel Accessories", "92pcs dremel accessories for dremel rotary tool"),
    ("Dowel cutter", "Dowel Cutter"),
    ("3 Jaw lathe chuck m12*45", "3 jaw lathe chuck M12*45"),
    ("Dowelling jig kikt", "Dowelling Jig Kit"),
    ("Electric Gun Nut Conversion Connector", "Electric Gun Nut Conversion Connector"),
    ("Quick position ruler", "Quick  Position Ruler"),
    ("Trimming machine", "Trimming Machine base ruler"),
    ("Drawer Handle Mounting Jig", "Drawer Handle Jig Locator"),
    ("71 pc Air Impact tool kit", "71pcs air impact tool"),
    ("300 pc drill bit set", "300 pcs Drill Bit Set"),
    ("pockrt hoe jig+clamps3", "Pocket Hole Jig + Clamps"),
    ("11pc hole saw kit", "11 pc Hole Saw Kit"),
    ("wood saw+blades", "Jig Wood Saw + Blade"),
    ("z-axis zero setter", "Z Axis Pre Setter"),
    ("drilling guide acccesories", "Drilling guide accesories"),
    ("chisel adapter (yellow)", "Chisel Adapter"),
    ("damageed bolt nuts remover", "Damaged Bolt Nuts Remover"),
    ("countersink bits with stopper", "Countersink Bits with Stopper"),
    ("mtaps m16", "M Taps (M16)"),
    ("mtaps m14", "M Taps (M14)"),
    ("mtaps m10", "M Taps (M10)"),
    ("mtaps m12", "M Taps (M12)"),
    ("ceramic cups(pink) s4", "Ceramic cups (4mm, pink)"),
    ("ceramic cups(pink) s5", "Ceramic cups (5mm, pink)"),
    ("ceramic cups(pink) s8", "Ceramic cups (8mm, pink)"),
    ("ceramic cups(pink) s6", "Ceramic cups (6mm, pink)"),
    ("ceramic cups(pink) p80", "Ceramic cups (p80, pink)"),
    ("ceramic cups(white) s4", "Ceramic cups (4mm, white)"),
    ("ceramic cups(white) s6", "Ceramic cups (6mm, white)"),
    ("ceramic cups(white) s8", "Ceramic cups (8mm, white)"),
    ("ceramic cups(white) s5", "Ceramic cups (5mm, white)"),
    ("ceramic cups(white) p80", "Ceramic cups (p80, white)"),
    ("4pc chisel set", "4pcs Chisel Set"),
    ("tungsten shank burr", "6mm Carbide Rotary Burrs"),
    ("glass and tile engraving bits", "2 in 1 Glass and Tile Cutter"),
    ("engraving bits", "5pc engraving bits"),
    ("center drillls (6pc)", "6pcs HSS Center Drills"),
    ("center drillls (10pc)", "10pc center drills"),
    ("air cinditioner tube expander", "5pc air conditioning tube expander"),
    ("inserts CCMT090304 VP15", "Carbide Inserts (CCMT090304 VP15)"),
    ("inserts MGMN300-MN3020", "Carbide Inserts (MGMN300-MN3020)"),
    ("inserts 11ER A60 BMA", "Carbide Inserts (11ER AG60 BMA)"),
    ("inserts 11IR A60 BMA", "Carbide Inserts (11IR AG60 BMA)"),
    ("inserts 16ER AG60 BMA", "Carbide Inserts (16ER AG60 BMA)"),
    ("inserts 16IR AG60 BMA", "Carbide Inserts (16IR AG60 BMA)"),
    ("inserts CCMT060204 VP15", "Carbide Inserts (CCMT060204 VP15)"),
    ("inserts DCMT070204 VP15", "Carbide Inserts (DCMT070204 VP15)"),
    ("inserts DCM11T304 VP15", "Carbide Inserts (DCMT11T304 VP15)"),
    ("inserts DNMG150604-TM 9125", "Carbide Inserts (DNMG150604-TM 9125)"),
    ("inserts TCMT110204 VP15", "Carbide Inserts (TCMT110204 VP15)"),
    ("inserts TCMT16T304VP15", "Carbide Inserts (TCMT16T304 VP15)"),
    ("inserts APMT1604PDER P3525", "Carbide Inserts (APMT1604PDER P3525)"),
    ("inserts MGMN400-M N3020", "Carbide Inserts (MGMN400-M N3020)"),
    ("inserts TNMG160408-TM 9125", "Carbide Inserts (TNMG160408-TM 9125)"),
    ("inserts MGMN200-G N3020", "Carbide Inserts (MGMN200-G N3020)"),
    ("inserts CNMG120408-HA H01", "Carbide Inserts (CNMG120408-HA H01)"),
    ("inserts APKT1135R08-MA H01", "Carbide Inserts (APKT1135R08- MAH01)"),
    ("inserts RPMT1204-JS", "Carbide Inserts (RPMT1204-JS)"),
    ("inserts TNMG160408-MS U6020", "Carbide Inserts (TNMG160408-MS U6020)"),
    ("1/2\" PIPE CLMP", "1/2\" pipe clamps"),
    ("3/4\" PIPE CLAMP WITHSTAND", "3/4\" Pipe Clamp With Stand"),
    ("WOODWORKING STEP DRILLS", "Wood step drill bits"),
    ("CHAMFERING TOOL SET", "Chamfering Tool For Angle Grinder"),
    ("CAPENTERS PEN + REFILL", "Capenter's Pencil With Graphite"),
    ("hinge hole jig 35 mm", "35mm hinge hole jig"),
    ("23 piece woodworking set incomplete", "23 pc Wood Working DIY Set"),
    ("radius plate", "Radius Plates"),
    ("abs angle ruler", "Abs angle ruler"),
    ("12 pc sockeet wrench", "12Pc socket wrench"),
    ("mt2 arvour", "MT2 B16 Arbor"),
    ("mt2-b16 arbour", "MT2-B16 Drill Chuck Arbor"),
    ("wireb stripper stripper", "Manual Wire stripper+blade"),
    ("micrometer screwgauge", "Micrometer screw gauge"),
    ("150 mm steel manual vernier calliper", "150MM Steel Manual Vernier Caliper"),
    ("rotary mini drill shaft", "Rotary mini drill shaft"),
    ("counterbore endmill", "7 pc counter bore endmill"),
    ("spring wheel divider", "Precision Spring Wing Divider"),
    ("Yellow price gun", "PRICE LABELLER GUN YELLOW"),
    ("price gun yellow", "PRICE LABELLER GUN YELLOW"),
    ("drill chuck arbour", "Drill Chuck Arbor MT3-B22"),
    ("water wash resin black", "SZ0001 Black Water Wash Resin"),
    ("glass bits", "10 pc Glass  Bits"),
    ("auger bits 8 pc complete", "Auger Bits  Set"),
    ("gypsum boar cutter", "Gypsum Board Cutting Tool"),
    ("vaccum sucking pen", "Vaccum sucking pen"),
    ("Allen Key", "8 pc Allen Key Set  (1.5 - 8mm)"),
    ("Distance measurer", "Distance Measurer CP-3010"),
    ("Plastic manual Vern ier caliper", "150MM Plastic Manual Vernier Caliper"),
    ("digital angle ruler", "300mm Digital Angle Ruler"),
    ("tile ruler adjustable punching hand tool", "Stainless steel tile ruler"),
    ("F Clamp 8\"", "F-clamp 8"),
    ("precision pocket ruler", "Precision pocket ruler"),
    ("multi gauge ruler", "Multigauge ruler"),
    ("360 degrees aluminium ruler", "360 degrees aluminium ruler"),
    ("150mm scriber", "150mm scriber"),
    ("T50 woodworking measuring tool", "T50 woodworking measuring tool"),
    ("hinge opening locator", "Hinge opening locator"),
    ("pipe tube cutter", "Tube Cutter"),
    ("heavy duty punching pliers", "Heavy Duty Punchipliers"),
    ("punching pliers (black)", "Punching Pliers"),
    ("6 pc flat wood bits", "6pc Flat Wood Bits"),
    ("5 pc jig saw blades", "5 pc Jig Saw Blade Set"),
    ("10 pc jig saw blades", "10 pc Jig Saw Blade Set"),
    ("pocket hole jig + clamps", "Pocket Hole Jig + Clamps"),
    ("dowel hole jig set", "3 in 1 Dowel Hole Jig set"),
    ("3 in 1 dowel hole jig set", "3 in 1 Dowel Hole Jig set"),
    ("countersink bits", "Countersink Bits"),
    ("countersink drill bits with stoper", "Countersink  Drill Bits with Stoper"),
    ("step drill bits (4-12mm)", "Step drill bits (4-12mm) (New)"),
    ("3 pc step drills", "3pc Step Drills"),
    ("3 pc HSS step drills", "3pc HSS Step Drills"),
    ("magnetic screw driver", "Magnetic Screw Driver"),
    ("telescoping gauge (6pc)", "Telescoping Gauge  (6pc)"),
    ("manual tap holder", "Manual Tap Holder"),
    ("m10 taps a pair", "M10 taps a pair"),
    ("m12 taps a pair", "M12 taps a pair"),
    ("m14 taps a pair", "M14 taps a pair"),
    ("m16 taps a pair", "M16 taps a pair"),
    ("wire stripper", "Wire Stripper"),
    ("plastic digital vernier caliper", "Plastic Digital Vernier Caliper"),
    ("150mm plastic manual vernier caliper", "150MM Plastic Manual Vernier Caliper"),
    ("300mm digital angle ruler", "300mm Digital Angle Ruler"),
    ("woodworking scribe ruler", "Woodworking scribe ruler"),
    ("yx-1 scribe gauge", "YX-1 Scribe Gauge"),
    ("wood working measuring tool kit", "Wood working Measuring tool kit"),
    ("angle grinder chamfer", "Angle Grinder Chamfer"),
    ("air sand blasting kit", "Air Sand Blasting kit"),
    ("air hydraulic riveter", "Air Hydraulic Riveter"),
    ("pneumatic rivet gun", "Pneumatic Rivet Gun"),
    ("rivet gun adapter", "Rivet Gun Adapter"),
    ("electric rivet adapter (black)", "Electric Rivet Adapter (black)"),
    ("electric rivet adapter (blue)", "Electric Rivet Adapter (blue)"),
    ("bottle cutter", "Bottle Cutter"),
    ("glass bottle cutter", "Glass Bottle Cutter"),
    ("furniture movers", "Furniture Movers"),
    ("heavy duty hole punch", "Heavy Duty Hole Punch"),
    ("M12*45mm wood chuck", "M12*45mm Wood Chuck"),
    ("er32 collets (3,2)", "ER-32 Collet (3mm)"),
    ("er32 collets (4,3)", "ER-32 Collet (4mm)"),
    ("er32 collets (5,4)", "ER-32 Collet (5mm)"),
    ("er32 collets (6,5)", "ER-32 Collet (6mm)"),
    ("er32 collets (7,6)", "ER-32 Collet (7mm)"),
    ("er32 collets (8,7)", "ER-32 Collet (8mm)"),
    ("er32 collets (9,8)", "ER-32 Collet (9mm)"),
    ("er32 collets (10,9)", "ER-32 Collet (10mm)"),
    ("er32 collets (11,10)", "ER-32 Collet (11mm)"),
    ("er32 collets (13,12)", "ER-32 Collet (13mm)"),
    ("er32 collets (14,13)", "ER-32 Collet (14mm)"),
    ("er32 collets (15,14)", "ER-32 Collet (15mm)"),
    ("er32 collets (16,15)", "ER-32 Collet (16mm)"),
    ("er32 collets (17,16)", "ER-32 Collet (17mm)"),
    ("er32 collets (18,17)", "ER-32 Collet (18mm)"),
    ("er32 collets (19,18)", "ER-32 Collet (19mm)"),
    ("er32 collets (20,19)", "ER-32 Collet (20mm)"),
    ("er16 collect (4,3)", "er16 Collet (4mm)"),
    ("er16 collect (5,4)", "er16 Collet (5mm)"),
    ("er16 collect (6,5)", "er16 Collet (6mm)"),
    ("er16 collect (7,6)", "er16 Collet (7mm)"),
    ("er16 collect (8,7)", "er16 Collet (8mm)"),
    ("endmills 4F D12*D12875L", "4 flute carbide endmill (12mm)"),
    ("endmills 4FD2*D4*50L", "4F*D2*D4*50 L"),
    ("endmills 4F D1*D4*50", "1mm 4 flute Carbide Endmill"),
    ("endmills 4F*D3*D4*50L", "4F*D3*D4*50 L"),
    ("endmills 4F*D8*D8*60L", "4 flute carbide endmill (8mm)"),
    ("endmills 4F*D10*D10*75L", None),
    ("endmills 4F*D6*D6*50L", "4 flute carbide endmill (6mm)"),
    ("endmills 4F*D5*D5*50L", "4 flute carbide endmill (5mm)"),
    ("endmills 4F*D4*D4*50L", "4 flute carbide endmill (4mm)"),
    ("endmills 4F*D2*D4*50L", "4 flute carbide endmill (2mm)"),
    ("endmills 4F*D3*D6*50L", "4 flute carbide endmill (3mm)"),

    # ---- Front2 specific ----
    ("SZ0007 Black ABS PRO2 Resin", "ABS Pro Resin Black"),
    ("SZ0008 Transparent ABS PRO2 Resin", "ABS Pro Resin Transparent"),
    ("SZ0009 Gray ABS PRO2 Resin", "ABS Pro Resin Grey"),
    ("11 mm Collet", "ER-32 Collet (11mm)"),
    ("13 mm Collet", "ER-32 Collet (13mm)"),
    ("14 mm Collet", "ER-32 Collet (14mm)"),
    ("15mm Collet", "ER-32 Collet (15mm)"),
    ("3mm Collet", "ER-32 Collet (3mm)"),
    ("2.0 Collet", "Tig collet (2.0mm)"),
    ("2.4 Collet", "Tig collet (2.4mm)"),
    ("K11 200 8inch 3jaw self centering", "lathe Chucks (self centering 3jaw, 8\")"),
    ("K1 2250 10inch 4 jaw independent", "lathe Chucks (Indepedent 4jaw, 10\")"),
    ("K12 CENTERING JAW 4JAW", "lathe Chucks (self centering 4jaw, 12\")"),
    ("K12 320 12inc 4 jaw self centered", "lathe Chucks (self centering 4jaw, 12\")"),
    ("Wood working table metre saw", "1200w Table Metr Saw"),
    ("HSS Square Twist Drill Bits 6mm", "4PCS HSS Twist Drill Bits"),
    ("HSS Square Twist Drill Bits 8mm", "4PCS HSS Twist Drill Bits"),
    ("HSS Square Twist Drill Bits 10mm", "4PCS HSS Twist Drill Bits"),
    ("HSS Square Twist Drill Bits 12mm", "4PCS HSS Twist Drill Bits"),
    ("Pipe Vise (2inch)", "Pipe Vise (max 2'')"),
    ("grinder angle stand black", "4.5 Angle Grinder stand (black)"),
    ("Counter bitts", "Countersink Bits"),
    ("Drill chucks (B16)", "Drill chucks (B16)"),
    ("Drill chucks (B22)", "Drill chucks (B22)"),
    ("ER 32 BIT 40", "BT40 -ER32-70"),
    ("Nylon peg clamp 6inch", "6\" Nylon Clamps"),
    ("Nylon peg clamp 3.inch", "3\" Nylon Clamps"),
    ("Nylon peg clamp 3.5.inch", "Nylon clamps 3.5"),
    ("Nylon peg clamp 2.5.inch", "2\" Nylon clamps"),
    ("3-5\" Set Nylon Clamps", "3-5\" Set Nylon Clamps"),
    ("1/4- 1 3/4 inch die stock", "1/4- 1 3/4 inch die stock"),
    ("1/2 -2 inch die stock", "1/2 -2 inch die stock"),
    ("32 PC Metric Tap and Die Set", "32 PC Metric Tap and Die Set"),
    ("12v Mini Die Grinder", "12V Mini Die Grinder"),
    ("7pc HSS Saw Blades", "7pc HSS Saw Blades"),
    ("6 pc Hss saw blades", "6 pc Hss saw blades"),
    ("10 pc HSS Drill Bits", "10pc HSS Drill Bits"),
    ("10pcs Cobalt Drill Bit Set", "10pcs Cobalt Drill Bit Set"),
    ("6pcs Cobalt Drill Bits", "6pcs Cobalt Drill Bits"),
    ("5pcs Counter Sink Drill Bits 2.8-4.0mm", "5pcs Counter Sink Drill Bits 2.8-4.0mm"),
    ("5pcs Tap Drill Bits", "5pcs Tap Drill Bits"),
    ("4pc 6-13mm HSS Tap Drills", "4pc 6-13mm HSS Tap Drills"),
    ("6pc HSS Drill Tap Bits", "6pc HSS Drill Tap Bits"),
    ("6pcs Assorted Files", "6pcs Assorted Files"),
    ("101 Pcs Drill Bit Set", "101 Pcs Drill Bit Set"),
    ("Hole Saw Set 15 20 25 30 35mm", "15 20 25 30 35mm Wood Hole Saw Set"),
    ("18pcs Set Hole Opener", "18pcs Set Hole Opener"),
    ("10pc Glass  Bits", "10 pc Glass  Bits"),
    ("5 pc Jig Saw Blade Set", "5 pc Jig Saw Blade Set"),
    ("10 pc Jig Saw Blade Set", "10 pc Jig Saw Blade Set"),
    ("7pc Mini Circular Saw Blades", "7pc Mini Circular Saw Blades"),
    ("20pc Mini Circular Saw Blades", "7pc Mini Circular Saw Blades"),
    ("Mini Circular Saw Blades", "7pc Mini Circular Saw Blades"),
    ("11 pc Hole Saw Kit", "11 pc Hole Saw Kit"),
    ("5pc Hole saw set", "5pc Hole saw set"),
    ("12Pc socket wrench", "12Pc socket wrench"),
    ("12Pc Socket wrench", "12Pc socket wrench"),
    ("46pcs car repair tool kit", "46pcs car repair tool kit"),
    ("53pc Car Repair Tool Set", "53pc Car Repair Tool Set"),
    ("71 Pcs Pneumatic Kit Set", "71 Pcs Pneumatic Kit Set"),
    ("41pc Air Impact tool kit", "41pc Air Impact tool kit"),
    ("45 pcs Air impact tool kit", "41pc Air Impact tool kit"),
    ("71pcs air impact tool", "71pcs air impact tool"),
    ("300 pcs Drill Bit Set", "300 pcs Drill Bit Set"),
    ("8 pc Allen Key Set  (1.5 - 8mm)", "8 pc Allen Key Set  (1.5 - 8mm)"),
    ("C Clamps/Drawer Installation", "C Clamps/Drawer Installation"),
    ("GH-101 AL Toggle Camps", "GH-101 AL Toggle Camps"),
    ("GH-201 BHB Toggle Clamps", "GH-201 BHB Toggle Clamps"),
    ("LHH2 clamp for T-7 track", "LHH2 clamp for T-7 track"),
    ("25LB  Welding Magnet", "25LB  Welding Magnet"),
    ("50lbs, 22kg Welding Magnets", "50lbs, 22kg Welding Magnets"),
    ("75lbs, 35kg Welding Magnets", "75lbs, 35kg Welding Magnets"),
    ("6pc Welding magnets", "6pc Welding magnets"),
    ("Welding Gloves", "Welding Gloves"),
    ("F-clamp 6\"", "F-clamp 6\""),
    ("F-clamp 8", "F-clamp 8"),
    ("10\" Bench Vise", "10\" Bench Vise"),
    ("8\" Bench Vise", "8\" Bench Vise"),
    ("8\" Bench Vise(200mm)", "8\" Bench Vise(200mm)"),
    ("bench vise with g-clamp", "bench vise with g-clamp"),
    ("Drill Holder", "Drill Holder"),
    ("Precise drill stand", "Portable Electric Drill Holder Stand"),
    ("Adjustable Circle Cutter", "Adjustable Circle Cutter"),
    ("Square Hole Locator", "Square Hole Locator"),
    ("Heavy Duty Hole Punch", "Heavy Duty Hole Punch"),
    ("Z Axis Pre Setter", "Z Axis Pre Setter"),
    ("Position Testing tool", "Position Testing tool"),
    ("Lever Arm Lifter", "Lever Arm Lifter"),
    ("Lever Arm Lifter Metal", "Lever Arm Lifter Metal"),
    ("Lever Hoist Blocks 3T*6M", "Lever Hoist Blocks 3T*6M"),
    ("Lever Hoist Blocks 6T*6M", "Lever Hoist Blocks 6T*6M"),
    ("HHY-300A Hydraulic Crimping Tool", "HHY-300A Hydraulic Crimping Tool"),
    ("HHYJ-240 Mechanical Crimping Tool", "HHYJ-240 Mechanical Crimping Tool"),
    ("HHYJ-50 Pressure Crimping Tool", "HHYJ-50 Pressure Crimping Tool"),
    ("Arc 315s Welding Machine", "ARC 315S Welding Machine"),
    ("MMA 250A Welding Machine", "MMA 250A Welding Machine"),
    ("NBC 270A 2in1 MIG welding Machine", "NBC 270A 2in1 MIG welding Machine"),
    ("200A 3M Mig Torch", "200A 3M Mig Torch"),
    ("350A 3M Mig Torch", "350A 3M Mig Torch"),
    ("30cm Gas Torch", "30cm Gas Torch"),
    ("35cm Gas Torch", "35cm Gas Torch"),
    ("30cm LPG Torch", "30cm LPG Torch"),
    ("35cm LPG Torch", "35cm LPG Torch"),
    ("50cm LPG Torch", "50cm LPG Torch"),
    ("40L CO2 CYLINDER WITH GAS", "40L CO2 CYLINDER WITH GAS"),
    ("15L CO2 Cylinder with Gas", "15L CO2 Cylinder with Gas"),
    ("Plasma Cutting Torch", "Plasma Cutting Torch"),
    ("Plasma Electrodes", "Plasma Electrodes (A40)"),
    ("Plasma Nozzles", "Plasma Nozzles (A40)"),
    ("Electrode Holder SH 800A", "Electrode Holder SH 800A"),
    ("Electrode Holder YX 500-800A", "Electrode Holder YX 500-800A"),
    ("Electrode holder", "Electrode Holder (SH 800A)"),
    ("Welding plugs (female, 10-25)", "Welding plugs (female, 10-25)"),
    ("Welding plugs (female, 35-50)", "Welding plugs (female, 35-50)"),
    ("Welding plugs (female, 50-70)", "Welding plugs (female, 50-70)"),
    ("Welding plugs (male, 10-25)", "Welding plugs (male, 10-25)"),
    ("Welding plugs (male, 35-50)", "Welding plugs (male, 35-50)"),
    ("Welding plugs (male, 50-70)", "Welding plugs (male, 50-70)"),
    ("Apron", "Apron"),
    ("3D Printing Pen Black", "3D Printing Pen Black"),
    ("3D Printing Pen Blue", "3D Printing Pen Blue"),
    ("3D Printing Pen Orange", "3D Printing Pen Orange"),
    ("3D Printing Pen Pink", "3D Printing Pen Pink"),
    ("3D Printing Pen Purple", "3D Printing Pen Purple"),
    ("3D Printing Pen Red", "3D Printing Pen Red"),
    ("3D Printing Pen Yellow", "3D Printing Pen Yellow"),
    ("3D Printing Nozzles (0.2mm)", "3D Printing Nozzles (0.2mm)"),
    ("3D Printing Nozzles (0.4mm)", "3D Printing Nozzles (0.4mm)"),
    ("Filament Dryer 1Spool", "Filament Dryer 1Spool"),
    ("Filament Dryer Box for 2 Spools", "Filament Dryer Box for 2 Spools"),
    ("Ender 3v3 plus", "Ender 3v3 plus"),
    ("Creality Ender-3 V3 3D Printer", "Creality Ender-3 V3 3D Printer (Model: SE)"),

    # ---- Consignment ----
    ("DOWELLING JIG KIT", "Dowelling Jig Kit"),
    ("DOWEL MAKER", "Dowel Cutter"),
    ("23 PC WOODWORKING SET", "23 pc Wood Working DIY Set"),
    ("15PC POCKET HOLE JIG KIT", "15Pcs Pocket Hole Jig 15 Degree Angle Oblique Hole Locator Dowel Drill Guide Kit"),
    ("144 PC MINI DRILL ACCESSORIES", "Mini Drill With Accesories(142)"),
    ("276 PCS MINIDRILL ACESSORIES", "276 Rotary Tool Kit"),
    ("3PC KIT TABLE PLATE", "3pcs Kit Table Plate"),
    ("ANGLE GRINDER CHAMFERING TOOL", "Chamfering Tool For Angle Grinder"),
    ("CREALITY ENDER 3V3 KE", "Creality Ender-3 V3 3D Printer (Model: KE)"),
    ("CO2 GAS REGULATOR", "CO2 Gas Regulator"),
    ("MATTE PLA+ DARK GREY(CREAT3D)", "Matte PLA Filament (Matte Grey PLA)"),
    ("MATTE PLA+ DARK BLUE(CREAT3D)", "Matte PLA Filament (Matte Blue PLA)"),
    ("MATTE PLA+ BLACK(CREAT3D)", "Matte PLA Filament (Matte Black PLA)"),
    ("MATTE PLA+ WHITE(CREAT3D)", "Matte PLA Filament (Matte White PLA)"),
    ("MATTE PLA+ PURPLE(CREAT3D)", "Matte PLA Filament (Matte Purple PLA)"),
    ("MATTE PLA+ COFFE(CREAT3D)", "Matte PLA Filament (Matte Shade Brown)"),
    ("MATTE PLA+ FAINT GREEN(CREAT3D)", "Matte PLA Filament (Matte Green PLA)"),
    ("MATTE PLA+ COCO BROWN(CREAT3D)", "Matte PLA Filament (Matte Brown PLA)"),
    ("MATTE PLA+ BEIGE(CREAT3D)", "Matte PLA Filament (Matte Skin PLA)"),
    ("MATTE PLA+ OLIVE GREEN(CREAT3D)", "Matte PLA Filament (Matte Green PLA)"),
    ("MATTE PLA+ ORANGE(CREAT3D)", "Matte PLA Filament (Matte Orange PLA)"),
    ("MATTE PLA+ GRASS GREEN(CREAT3D)", "Matte PLA Filament (Matte Grass Green PLA)"),
    ("MATTE PLA+ YELLOW(CREAT3D)", "Matte PLA Filament (Matte Yellow PLA)"),
    ("MATTE PLA+ RED(CREAT3D)", "Matte PLA Filament (Matte Red PLA)"),
    ("MATTE PLA+ SILICON GREEN(CREAT3D)", "Matte PLA Filament (Matte Green PLA)"),
    ("MATTE PLA+ TANGERINE(CREAT3D)", "Matte PLA Filament (Matte Orange PLA)"),
    ("MATTE PLA+ SKY BLUE(CREAT3D)", "Matte PLA Filament (Matte Blue PLA)"),
    ("MATTE PLA+ WINTER WHITE(CREAT3D)", "Matte PLA Filament (Matte White PLA)"),
    ("MATTE PLA+ PINK(CREAT3D)", "Matte PLA Filament (Matte Lillac Purple PLA)"),
    ("PLA BASIC BRONZE(CREAT3D)", "PLA Basic (Golden PLA Basic)"),
    ("PLA BASIC RED MAROON(CREAT3D)", "PLA Basic (Red PLA Basic)"),
    ("PLA+ BLACK(CREAT3D)", "PLA + (Black PLA +)"),
    ("PLA+ WHITE(CREAT3D)", "PLA + (White PLA +)"),
    ("POLYMAKER MATTE EARTH BROWN", "Matte PLA Filament (Matte Brown PLA)"),
    ("POLYMAKER MATTE WOOD BROWN", "Matte PLA Filament (Matte Brown PLA)"),
    ("POLYMAKER MATTE CHARCOAL BLACK", "Matte PLA Filament (Matte Black PLA)"),
    ("POLYMAKER MATTE COTTON WHITE", "polymakers (cotton white)"),
    ("PETG WHITE(CREAT3D)", "PETG (White PETG, Spool Weight: 1Kg)"),
    ("PETG BLACK(CREAT3D)", "PETG (Black PETG, Spool Weight: 1Kg)"),

    # ---- FROM-WAREHOUSE ----
    ("Polymaker matte cotton white", "polymakers (cotton white)"),
    ("PETG white", "PETG (White PETG, Spool Weight: 1Kg)"),
    ("PETG white(kingroon)", "PETG (White PETG, Spool Weight: 1Kg)"),
    ("PLA Black", "PLA Basic (Black PLA Basic)"),
    ("300mm Vernier calipers", "300 mm vanier calliper"),
    ("PLA Blue", "PLA Basic (Blue PLA)"),
    ("PLA Skyblue", "PLA Basic (Sky Blue PLA Basic)"),
    ("PETG Black", "PETG (Black PETG, Spool Weight: 1Kg)"),
    ("PETG Red", "PETG (Red PETG, Spool Weight: 1Kg)"),
    ("PETG Skyblue", "PETG (Sky Blue PETG, Spool Weight: 1Kg)"),
    ("PLA Dark Grey", "PLA Basic (Dark Grey PLA Basic)"),
    ("Matte Black", "Matte PLA Filament (Matte Black PLA)"),
    ("Matte Shade Brown", "Matte PLA Filament (Matte Shade Brown)"),
    ("PLA Matte Black", "Matte PLA Filament (Matte Black PLA)"),
    ("PLA Matte White", "Matte PLA Filament (Matte White PLA)"),
    ("TPU Black", "TPU Filaments (Black TPU)"),
    ("Argon Regulators (female)", "Argon regulator female"),
    ("PETG Dark Blue(kingroon)", "PETG (Dark Blue PETG, Spool Weight: 1Kg)"),
    ("PETG Transparent(kingroon)", "PETG (Transparent PETG, Spool Weight: 1Kg)"),
    ("PLA White(kingroon)", "PLA Basic (White PLA Basic)"),
    ("PLA Black(kingroon)", "PLA Basic (Black PLA Basic)"),
    ("3/4\" pipe clamps", "3/4'' Pipe Clamp without stand"),
    ("PLA+ Black (Antisky)", "PLA + (Black PLA +)"),
    ("PLA+ White(Antisky)", "PLA + (White PLA +)"),
    ("Angle Grinder Stand (green)", "Angle Grinder Stand with Sliding Mechanism"),

    # ---- PI shipment KR2026042701 (direct to Ruiru) ----
    # Yellow-highlighted rows (items 31-38: Nylon/ASA/ABS) did NOT arrive
    # and are excluded in parse_pi(); the lines below are the ones included.
    ("NPLA001 PLA Filament 1kg(White)", "PLA Basic (White PLA Basic)"),
    ("NPLA002 PLA Filament 1kg(Black)", "PLA Basic (Black PLA Basic)"),
    ("NPLA006 PLA Filament 1kg(Blue)", "PLA Basic (Blue PLA)"),
    ("NPLA019 PLA Filament 1kg(Sky Blue)", "PLA Basic (Sky Blue PLA Basic)"),
    ("NPLA018 PLA Filament 1kg(Dark Grey)", "PLA Basic (Dark Grey PLA Basic)"),
    ("NPLA010 PLA Filament 1kg(Red)", "PLA Basic (Red PLA Basic)"),
    ("KR-Silk003Y-1CH Silk Pla 1kg (Black)", "Silk PLA (Silk Black PLA)"),
    ("KR-Silk001Y-1CH Silk Pla 1kg (Golden)", "Silk PLA (Silk Golden)"),
    ("KR-PLA301Y-1CH matte pla (black)", "Matte PLA Filament (Matte Black PLA)"),
    ("KR-PLA302Y-1CH matte pla (white)", "Matte PLA Filament (Matte White PLA)"),
    ("KR-PLA312Y-1CH matte pla (Shade brown)", "Matte PLA Filament (Matte Shade Brown)"),
    ("NPETG087 Petg Filament 1kg(Black)", "PETG (Black PETG, Spool Weight: 1Kg)"),
    ("NPETG088 Petg Filament 1kg(White)", "PETG (White PETG, Spool Weight: 1Kg)"),
    ("NPETG006 Petg Filament 1kg(Gray)", "PETG (Grey PETG, Spool Weight: 1Kg)"),
    ("NPETG001 Petg Filament 1kg(Red)", "PETG (Red PETG, Spool Weight: 1Kg)"),
    ("NPETG002 Petg Filament 1kg(Sky Blue)", "PETG (Sky Blue PETG, Spool Weight: 1Kg)"),
    ("NPETG003 Petg Filament 1kg(Yellow)", "PETG (Yellow PETG, Spool Weight: 1Kg)"),
    ("NPETG007 Petg Filament 1kg(Transaprent)", "PETG (Transparent PETG, Spool Weight: 1Kg)"),
    ("NPETG018 Petg Filament 1kg(Green)", "PETG (Green PETG, Spool Weight: 1Kg)"),
    ("NPETG017 Petg Filament 1kg(Orange)", "PETG (Orange PETG, Spool Weight: 1Kg)"),
    ("NPETG019 Petg Filament 1kg(Dark Blue)", "PETG (Dark Blue PETG, Spool Weight: 1Kg)"),
    ("KR-TPU001Y-1CH Black", "TPU Filaments (Black TPU)"),
    ("KR-TPU002Y-1CH White", "TPU Filaments (White TPU)"),
    ("KR-TPU004Y-1CH Transparent", "TPU Filaments (Transparent TPU)"),
    ("KR-TPU006Y-1CH Gray", "TPU Filaments (Grey TPU)"),
    ("KR-TPU003Y-1CH Red", "TPU Filaments (Red TPU)"),
    ("KR-TPU008Y-1CH Blue", "TPU Filaments (Blue TPU)"),
    ("KR-TPU009Y-1CH Orange", "TPU Filaments (Orange TPU)"),
    ("KR-TPU005Y-1CH Green", "TPU Filaments (Green TPU)"),
    ("KR-TPU007Y-1CH Yellow", "TPU Filaments (Yellow TPU)"),
    ("Basic Pla 3KG spool (white )", "PLA Basic (White) 3kg"),
    ("Basic Pla 3KG spool (Black )", "PLA Basic (Black) 3kg"),
    ("XP241001 Filament Dryer Box for 2 Spools", "Filament Dryer Box for 2 Spools"),
    ("XP250506 Filament Dryer", "Filament Dryer 1Spool"),

    # ---- Negative overrides: known NOT in system ----
    ("Orbit Sander", None),
    ("Welding Overall", None),
    ("FFQ 9939", None),
    ("46 Socket 1/4", None),
    ("6inch bench vise", None),
    ("4inch bench vise", None),
    ("F-clamp 4", None),
    ("C Clamp 7", None),
    ("C Clamp 11/14/18", None),
    ("2\" G-CLAMP", None),
    ("Table tool cross cylder", None),
    ("Drill bit sharpener", None),
    ("Welding tool kit bag", None),
    ("Lithium Iron Battery", None),
    ("Lithium Iron drawer(faulty)", None),
    ("P120 sand paper", None),
    ("P180 Sand paper", None),
    ("Grinding stone", None),
    ("Hand telescopic quick suppot rod", None),
    ("ceramic hole locater", None),
    ("Precise electric green drill", None),
    ("Mini drill without accessories", None),
    ("Tick torch Accessories", None),
    ("Plastcic jig", None),
    ("ER 20 K11 Collet", None),
    ("K7200 4 JAW Indipendent", None),
    ("K12 CENTERING JAW", None),
    ("Table tool cross cylder", None),
    ("100 mm Vanier Calliper", None),
    ("Aluminium Angle", None),
    ("HSS Drill Bitts set", None),
    ("Router Locator", None),
    ("53pc socket set 1/4 dr", None),
]


def build_overrides(variants):
    """Resolve the override table against the live variant list.
    Returns dict {norm(manual): product_id or None}."""
    by_disp = {}
    for v in variants:
        disp = (v.get("display_name") or v.get("name") or "").strip()
        by_disp[disp_key(disp)] = v["id"]

    overrides = {}
    unresolved = []
    for manual, target in _OVERRIDE_TABLE:
        key = match_key(manual)
        if target is None:
            overrides[key] = None
            continue
        tid = by_disp.get(disp_key(target))
        if tid is None:
            unresolved.append((manual, target))
            continue
        overrides[key] = tid
    if unresolved:
        print("WARNING: override targets not found in system:")
        for m, t in unresolved:
            print(f"  {m!r} -> {t!r}")
    return overrides


def build_report(system, sales_lines, overrides):
    variants = build_variant_objects(system["variants"], system["templates"])
    matcher = Matcher(variants, overrides)
    nbi_id, rui_id = system["nbi_id"], system["rui_id"]
    loc_stock = system["loc_stock"]

    # current system stock per product per location
    cur_nbi = defaultdict(float)
    cur_rui = defaultdict(float)
    for pid, locs in loc_stock.items():
        cur_nbi[pid] = locs.get(nbi_id, 0)
        cur_rui[pid] = locs.get(rui_id, 0)

    # sales per product (units) since 06-07
    sold = defaultdict(float)
    for l in sales_lines:
        sold[l["product_id"]] += l["qty"]

    # physical counts
    f1 = aggregate(parse_front1())
    f2 = aggregate(parse_front2())
    cons = aggregate(parse_consignment())
    fw = aggregate(parse_from_warehouse())
    wh = aggregate(parse_warehouse())
    pi = aggregate(parse_pi())

    count_data = {
        "front1": f1, "front2": f2, "consignment": cons,
        "from_warehouse": fw, "warehouse": wh, "pi": pi,
    }

    match_log = []  # manual-name, qty, pid, score, source
    by_pid = defaultdict(lambda: {"count": 0.0, "consignment": 0.0,
                                  "from_wh": 0.0, "warehouse": 0.0,
                                  "pi": 0.0,
                                  "sources": set(), "names": []})

    def apply(source_name, aggmap, bucket):
        for name, qty in aggmap.items():
            pid, score = matcher.match(name)
            entry = {"name": name, "qty": qty, "pid": pid, "score": score, "source": source_name}
            match_log.append(entry)
            if pid is not None:
                b = by_pid[pid]
                b[bucket] += qty
                b["sources"].add(source_name)
                b["names"].append(f"{name} ({qty:g})")

    apply("Front 1", f1, "count")
    apply("Front 2", f2, "count")
    apply("Consignment", cons, "consignment")
    apply("FROM-WAREHOUSE", fw, "from_wh")
    apply("Warehouse", wh, "warehouse")
    apply("PI", pi, "pi")

    # products counted that are not in system
    unmatched = {}
    for m in match_log:
        if m["pid"] is None:
            key = (m["name"], m["source"])
            unmatched.setdefault(key, 0.0)
            unmatched[key] += m["qty"]

    return {
        "variants": variants, "matcher": matcher, "by_pid": by_pid,
        "match_log": match_log, "count_data": count_data, "unmatched": unmatched,
        "cur_nbi": cur_nbi, "cur_rui": cur_rui, "sold": sold,
        "nbi_id": nbi_id, "rui_id": rui_id,
    }


# ---------------------------------------------------------
# EXCEL WRITERS
# ---------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="E2E8F0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_sheet(ws, widths=None, freeze="A2"):
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
    if widths:
        for letter, w in widths.items():
            ws.column_dimensions[letter].width = w
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions


def write_merged_count(report):
    rows = []
    for m in report["match_log"]:
        rows.append({
            "Manual Item": m["name"],
            "Source": m["source"],
            "Qty": m["qty"],
            "Matched System Product": report["variants_by_id"].get(m["pid"], {}).get("display") if m["pid"] else "(not in system)",
            "Match Confidence": m["score"],
        })
    df = pd.DataFrame(rows)
    path = os.path.join(DOCS, "NAIROBI_COUNT_MERGED_FRONT1_FRONT2.xlsx")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Merged Count")
        style_sheet(writer.sheets["Merged Count"],
                    widths={"A": 38, "B": 14, "C": 8, "D": 48, "E": 16})
    return path


def write_sales_report(sales_lines, start="2026-07-06"):
    rows = [{
        "Date": l["date"], "Channel": l["channel"], "Order": l["order_id"],
        "Product": l["product_name"], "Category": l["category"],
        "Qty": l["qty"], "Revenue": round(l["revenue"], 2),
    } for l in sales_lines]
    end = max((l["date"] for l in sales_lines), default=start)
    path = os.path.join(DOCS, f"SALES_{start.replace('-', '')}_to_{end.replace('-', '')}.xlsx")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df = pd.DataFrame(rows)
        df.to_excel(writer, index=False, sheet_name="Sales Lines")
        style_sheet(writer.sheets["Sales Lines"],
                    widths={"A": 12, "B": 14, "C": 10, "D": 48, "E": 22, "F": 8, "G": 12})

        summary = df.groupby(["Product", "Category"], as_index=False).agg(
            Qty=("Qty", "sum"), Revenue=("Revenue", "sum"))
        summary = summary.sort_values("Qty", ascending=False)
        summary.to_excel(writer, index=False, sheet_name="Summary by Product")
        style_sheet(writer.sheets["Summary by Product"],
                    widths={"A": 48, "B": 22, "C": 8, "D": 12})

        byday = df.groupby(["Date", "Channel"], as_index=False).agg(
            Qty=("Qty", "sum"), Revenue=("Revenue", "sum"))
        byday = byday.sort_values("Date")
        byday.to_excel(writer, index=False, sheet_name="Daily Totals")
        style_sheet(writer.sheets["Daily Totals"],
                    widths={"A": 12, "B": 14, "C": 8, "D": 12})
    return path


def write_final_report(report):
    vid_map = report["variants_by_id"]
    by_pid = report["by_pid"]
    cur_nbi, cur_rui = report["cur_nbi"], report["cur_rui"]
    sold = report["sold"]
    rows = []

    def num(v):
        return 0.0 if v is None else float(v)

    pids = set(by_pid.keys()) | set(cur_nbi.keys()) | set(cur_rui.keys()) | set(sold.keys())
    for pid in sorted(pids):
        v = vid_map.get(pid)
        if v is None:
            continue
        b = by_pid[pid]
        counted = num(b["count"])
        cons = num(b["consignment"])
        fw = num(b["from_wh"])
        wh = num(b["warehouse"])
        pi = num(b["pi"])
        cur_n = num(cur_nbi.get(pid))
        cur_r = num(cur_rui.get(pid))
        s = num(sold.get(pid))
        exp_n = (counted if counted else cur_n) + cons + fw - s
        exp_r = (wh if wh else cur_r) - fw + pi
        status = []
        if counted:
            status.append("Physical count")
        if cons:
            status.append("Consignment")
        if fw:
            status.append("From warehouse")
        if wh:
            status.append("Warehouse count")
        if pi:
            status.append("PI shipment")
        if exp_n < 0:
            status.append("SHORTFALL: NBI")
        if exp_r < 0:
            status.append("SHORTFALL: RUI")
        rows.append({
            "Product": v["display"], "Category": v["template"],
            "Current NBI (System)": cur_n, "Current RUI (System)": cur_r,
            "Physical Count NBI": counted, "Consignment Added": cons,
            "From Warehouse (+NBI/-RUI)": fw, "Warehouse Count RUI": wh,
            "PI Shipment (RUI)": pi,
            "Sold since 06-Jul": s,
            "Expected NBI": exp_n, "Expected RUI": exp_r,
            "Status": ", ".join(status) or "-",
        })

    # items found physically but not in the system
    for (name, source), qty in sorted(report["unmatched"].items()):
        rows.append({
            "Product": f"{name}  [{source}]", "Category": "(not in system)",
            "Current NBI (System)": "", "Current RUI (System)": "",
            "Physical Count NBI": qty, "Consignment Added": "",
            "From Warehouse (+NBI/-RUI)": "", "Warehouse Count RUI": "",
            "Sold since 06-Jul": "", "Expected NBI": qty, "Expected RUI": "",
            "Status": "Not in system - verify",
        })

    df = pd.DataFrame(rows)
    path = os.path.join(DOCS, "EXPECTED_STOCK.xlsx")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Expected Stock")
        style_sheet(writer.sheets["Expected Stock"],
                    widths={"A": 52, "B": 24, "C": 14, "D": 14, "E": 14, "F": 14,
                            "G": 18, "H": 14, "I": 14, "J": 12, "K": 12, "L": 12, "M": 26})
    return path


def main():
    print("Fetching system data...")
    system = fetch_system()
    print("Fetching sales...")
    sales_lines = fetch_sales("2026-07-06")

    overrides = build_overrides(system["variants"])

    report = build_report(system, sales_lines, overrides)
    vid_map = {v["id"]: v for v in report["variants"]}
    report["variants_by_id"] = vid_map

    print("Sales lines:", len(sales_lines))
    print("Front1 items:", len(report["count_data"]["front1"]),
          "Front2 items:", len(report["count_data"]["front2"]),
          "Consignment:", len(report["count_data"]["consignment"]),
          "From-warehouse:", len(report["count_data"]["from_warehouse"]),
          "Warehouse items:", len(report["count_data"]["warehouse"]),
          "PI items:", len(report["count_data"]["pi"]))

    nm = sum(1 for m in report["match_log"] if m["pid"] is None)
    low = sum(1 for m in report["match_log"] if m["score"].startswith("fuzzy") and float(m["score"].split(":")[1]) < 92)
    print(f"NOMATCH items: {nm}   low-confidence fuzzy: {low}")

    # Write merged count
    p = write_merged_count(report)
    print("Merged count written:", p)

    # Write sales report
    p = write_sales_report(sales_lines)
    print("Sales report written:", p)

    # Write final expected stock
    p = write_final_report(report)
    print("Final expected stock written:", p)

    # Dump match log for review
    with open(os.path.join(BASE, "_match_review.tsv"), "w", encoding="utf-8") as f:
        f.write("SOURCE\tMANUAL\tQTY\tMATCHED\tPID\tSCORE\n")
        for m in report["match_log"]:
            matched = vid_map.get(m["pid"], {}).get("display", "") if m["pid"] else "(NOMATCH)"
            f.write(f"{m['source']}\t{m['name']}\t{m['qty']:g}\t{matched}\t{m['pid']}\t{m['score']}\n")
    print("Match review written: _match_review.tsv")


if __name__ == "__main__":
    main()